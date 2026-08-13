"""Evaluate prostate segmentation results and export an Eval—sum-style workbook.

The output layout follows ``D:\\WUSI\\Prostate\\Eval—sum.xlsx``.  In addition
to overlap, distance, confusion-matrix and volume metrics, it reports four
slice statistics:

* Ref_slice: number of axial slices containing a reference mask.
* Pred_slice: number of axial slices containing a predicted mask.
* overseg_slice: predicted-positive but reference-empty slices.
* underseg_slice: reference-positive but predicted-empty slices.
"""

from __future__ import annotations

import os
import re
import subprocess
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import SimpleITK as sitk
from medpy import metric
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


COLUMNS = [
    "ID", "2D Dice", "2D HD95 (mm)",
    "3D Dice", "3D HD95 (mm)", "IoU", "ASD (mm)",
    "Sensitivity (%)", "Specificity (%)", "Precision (%)", "Accuracy (%)",
    "TP (vox)", "TN (vox)", "FP (vox)", "FN (vox)",
    "Ref_volume (cc)", "Pred_volume (cc)",
    "Ref_slice", "Pred_slice", "overseg_slice", "underseg_slice",
]

MEAN_STD_COLUMNS = {
    "2D Dice", "2D HD95 (mm)", "3D Dice", "3D HD95 (mm)", "IoU",
    "ASD (mm)", "Ref_volume (cc)", "Pred_volume (cc)",
    "Ref_slice", "Pred_slice", "overseg_slice", "underseg_slice",
}


def dice_coefficient(a: np.ndarray, b: np.ndarray) -> float:
    smooth = 1e-5
    a = a > 0.5
    b = b > 0.5
    intersection = np.logical_and(a, b).sum()
    return float((2.0 * intersection + smooth) / (a.sum() + b.sum() + smooth))


def compute_2d_metrics(
    gt_array: np.ndarray,
    pred_array: np.ndarray,
    in_plane_spacing: tuple[float, float],
) -> tuple[float, float]:
    """Compute slice-wise Dice/HD95 on reference-positive axial slices."""
    dice_values: list[float] = []
    hd95_values: list[float] = []
    # NumPy axes are (z, y, x), hence spacing is (y, x) here.
    spacing_yx = (in_plane_spacing[1], in_plane_spacing[0])

    for gt_slice, pred_slice in zip(gt_array, pred_array):
        if not gt_slice.any():
            continue
        dice_values.append(dice_coefficient(pred_slice, gt_slice))
        if pred_slice.any():
            try:
                hd95_values.append(
                    float(metric.binary.hd95(pred_slice, gt_slice, voxelspacing=spacing_yx))
                )
            except Exception:
                pass

    dice_2d = round(float(np.mean(dice_values)), 2) if dice_values else 0.0
    hd95_2d = round(float(np.mean(hd95_values)), 2) if hd95_values else 0.0
    return dice_2d, hd95_2d


def compute_slice_statistics(
    gt_array: np.ndarray, pred_array: np.ndarray
) -> tuple[int, int, int, int]:
    gt_positive = np.any(gt_array, axis=(1, 2))
    pred_positive = np.any(pred_array, axis=(1, 2))
    ref_slice = int(gt_positive.sum())
    pred_slice = int(pred_positive.sum())
    overseg_slice = int(np.logical_and(pred_positive, ~gt_positive).sum())
    underseg_slice = int(np.logical_and(gt_positive, ~pred_positive).sum())
    return ref_slice, pred_slice, overseg_slice, underseg_slice


def run_plastimatch(gt_path: Path, pred_path: Path) -> tuple[float, ...]:
    result = subprocess.run(
        ["plastimatch", "dice", str(gt_path), str(pred_path)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=True,
    )

    def extract(pattern: str) -> float:
        match = re.search(pattern, result.stdout)
        return float(match.group(1)) if match else 0.0

    dice_3d = round(extract(r"DICE:\s*([0-9.eE+-]+)"), 2)
    sensitivity = round(extract(r"SE:\s*([0-9.eE+-]+)") * 100, 2)
    specificity = round(extract(r"SP:\s*([0-9.eE+-]+)") * 100, 2)
    tp = int(extract(r"TP:\s*([0-9.eE+-]+)"))
    tn = int(extract(r"TN:\s*([0-9.eE+-]+)"))
    fp = int(extract(r"FP:\s*([0-9.eE+-]+)"))
    fn = int(extract(r"FN:\s*([0-9.eE+-]+)"))

    hd_result = subprocess.run(
        ["plastimatch", "dice", "--hausdorff", str(gt_path), str(pred_path)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=True,
    )
    hd_match = re.search(
        r"Percent\s*\(0\.95\)\s*Hausdorff\s*distance\s*\(boundary\)\s*=\s*([0-9.eE+-]+)",
        hd_result.stdout,
    )
    hd95_3d = round(float(hd_match.group(1)), 2) if hd_match else np.nan
    return dice_3d, hd95_3d, sensitivity, specificity, tp, tn, fp, fn


def evaluate_case(gt_path: Path, pred_path: Path) -> list[float | int]:
    gt_img = sitk.ReadImage(str(gt_path))
    pred_img = sitk.ReadImage(str(pred_path))
    if gt_img.GetSize() != pred_img.GetSize():
        raise ValueError(f"image size mismatch: GT={gt_img.GetSize()}, pred={pred_img.GetSize()}")

    gt_array = sitk.GetArrayFromImage(gt_img) > 0
    pred_array = sitk.GetArrayFromImage(pred_img) > 0
    spacing = gt_img.GetSpacing()

    dice_2d, hd95_2d = compute_2d_metrics(gt_array, pred_array, spacing[:2])
    dice_3d, hd95_3d, se, sp, tp, tn, fp, fn = run_plastimatch(gt_path, pred_path)

    iou = round(tp / (tp + fp + fn), 2) if tp + fp + fn else 0.0
    precision = round(tp / (tp + fp) * 100, 2) if tp + fp else 0.0
    accuracy = round((tp + tn) / (tp + tn + fp + fn) * 100, 2) if tp + tn + fp + fn else 0.0

    try:
        asd = round(
            float(metric.binary.assd(pred_array, gt_array, voxelspacing=spacing[::-1])), 2
        )
    except Exception:
        asd = 0.0

    voxel_volume = float(np.prod(spacing))
    ref_volume = round(float(gt_array.sum()) * voxel_volume / 1000, 2)
    pred_volume = round(float(pred_array.sum()) * voxel_volume / 1000, 2)
    slice_stats = compute_slice_statistics(gt_array, pred_array)

    return [
        dice_2d, hd95_2d, dice_3d, hd95_3d, iou, asd,
        se, sp, precision, accuracy, tp, tn, fp, fn,
        ref_volume, pred_volume, *slice_stats,
    ]


def add_statistics_rows(df: pd.DataFrame) -> pd.DataFrame:
    patient_rows = df[df["ID"].astype(str).str.fullmatch(r"p_\d+", na=False)].copy()
    if patient_rows.empty:
        return df

    mean_row: dict[str, object] = {"ID": "Mean"}
    std_row: dict[str, object] = {"ID": "STD"}
    for column in COLUMNS[1:]:
        values = pd.to_numeric(patient_rows[column], errors="coerce")
        mean_row[column] = round(float(values.mean()), 2) if values.notna().any() else ""
        std_row[column] = (
            round(float(values.std(ddof=1)), 2)
            if column in MEAN_STD_COLUMNS and values.notna().sum() > 1
            else ""
        )
    return pd.concat([df, pd.DataFrame([mean_row, std_row])], ignore_index=True)


def format_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for column_index in range(1, ws.max_column + 1):
        letter = get_column_letter(column_index)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in ws[letter])
        ws.column_dimensions[letter].width = min(max_length + 2, 40)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B2"


def evaluate_model(gt_dir: str, pred_dir: str, sheet_name: str, output_excel: str) -> None:
    print(f"\nEvaluating model: {sheet_name}")
    records: list[list[object]] = []
    pred_root = Path(pred_dir)
    gt_root = Path(gt_dir)

    for pred_path in sorted(pred_root.glob("*.nii.gz")):
        stem = pred_path.name.removesuffix(".nii.gz")
        number_match = re.search(r"_(\d+)$", stem)
        case_id = f"p_{int(number_match.group(1))}" if number_match else stem
        gt_path = gt_root / pred_path.name
        if not gt_path.exists():
            print(f"Missing GT: {gt_path}")
            records.append([case_id] + [""] * (len(COLUMNS) - 1))
            continue
        try:
            records.append([case_id, *evaluate_case(gt_path, pred_path)])
        except Exception as exc:
            print(f"Failed {case_id}: {exc}")
            records.append([case_id] + [""] * (len(COLUMNS) - 1))

    df = add_statistics_rows(pd.DataFrame(records, columns=COLUMNS))
    mode = "a" if os.path.exists(output_excel) else "w"
    writer_args = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_args["if_sheet_exists"] = "replace"
    with pd.ExcelWriter(output_excel, **writer_args) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    workbook = openpyxl.load_workbook(output_excel)
    format_sheet(workbook[sheet_name])
    workbook.save(output_excel)


def generate_summary_sheet(output_excel: str) -> None:
    workbook = openpyxl.load_workbook(output_excel, data_only=True)
    rows: list[list[object]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name == "Summary_Mean":
            continue
        ws = workbook[sheet_name]
        headers = [cell.value for cell in ws[1]]
        mean_values = next(
            ([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
             for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == "Mean"),
            None,
        )
        std_values = next(
            ([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
             for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == "STD"),
            None,
        )
        if mean_values is None or std_values is None:
            continue
        summary_row: list[object] = [sheet_name]
        for index, column in enumerate(headers[1:], start=1):
            mean_value, std_value = mean_values[index], std_values[index]
            if column in MEAN_STD_COLUMNS and mean_value != "" and std_value != "":
                summary_row.append(f"{float(mean_value):.2f} ± {float(std_value):.2f}")
            elif isinstance(mean_value, (int, float)):
                summary_row.append(round(float(mean_value), 2))
            else:
                summary_row.append("")
        rows.append(summary_row)

    columns = ["模型名称", *COLUMNS[1:]]
    summary_df = pd.DataFrame(rows, columns=columns)
    workbook = openpyxl.load_workbook(output_excel)
    if "Summary_Mean" in workbook.sheetnames:
        workbook.remove(workbook["Summary_Mean"])
    ws = workbook.create_sheet("Summary_Mean", 0)
    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(row)
    format_sheet(ws)
    workbook.save(output_excel)


def main() -> None:
    # Adjust these paths for the experiment to evaluate.
    gt_dir = r"D:\A-project\Prostate\nnUNet\TestResults_Seminal\CTV1\labelsTs"
    output_excel = r"D:\A-project\Prostate\nnUNet\TestResults_Seminal\CTV1\Eval—sum.xlsx"
    model_paths = {
        "No": r"D:\A-project\Prostate\nnUNet\TestResults_Seminal\CTV1\TestResults_category\no",
        "Yes": r"D:\A-project\Prostate\nnUNet\TestResults_Seminal\CTV1\TestResults_category\yes",
    }

    Path(output_excel).parent.mkdir(parents=True, exist_ok=True)
    for name, path in model_paths.items():
        evaluate_model(gt_dir, path, name, output_excel)
    generate_summary_sheet(output_excel)
    print(f"\nEvaluation completed: {output_excel}")


if __name__ == "__main__":
    main()
