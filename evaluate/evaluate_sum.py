"""
2d：仅GT切片
"""

import os
import re
import numpy as np
import pandas as pd
import subprocess
import SimpleITK as sitk
from medpy import metric
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font


# ==========================================================
# -------------------- 2D 评估函数 --------------------------
# ==========================================================

def DiceCoefficient(a, b):
    SMOOTH = 1e-5
    a = a > 0.5
    b = b > 0.5
    intersection = np.sum(np.logical_and(a, b))
    return (2. * intersection + SMOOTH) / (np.sum(a) + np.sum(b) + SMOOTH)


def ComputeDSC2D(gtDir, predDir):
    gtVolume = sitk.GetArrayFromImage(sitk.ReadImage(gtDir))
    predVolume = sitk.GetArrayFromImage(sitk.ReadImage(predDir))
    dice, count = 0.0, 0
    for i in range(predVolume.shape[0]):
        gtMat = (gtVolume[i] > 0).astype(np.uint8)
        predMat = (predVolume[i] > 0).astype(np.uint8)
        # gt与pred均非空的切片
        # if np.max(gtMat) == 0 or np.max(predMat) == 0:
        # 仅gt切片
        if np.max(gtMat) == 0 :
            continue

        dice += DiceCoefficient(predMat, gtMat)
        count += 1
    return round(dice / count, 2) if count > 0 else 0


def ComputeHD95(gtDir, predDir):
    oridata = sitk.ReadImage(gtDir)
    fColSpacing = oridata.GetSpacing()[1]
    gtVolume = sitk.GetArrayFromImage(oridata)
    predVolume = sitk.GetArrayFromImage(sitk.ReadImage(predDir))
    hd_list = []
    for i in range(predVolume.shape[0]):
        gtMat = (gtVolume[i] > 0).astype(np.uint8)
        predMat = (predVolume[i] > 0).astype(np.uint8)
        # gt与pred均非空的切片
        # if np.max(gtMat) > 0 and np.max(predMat) > 0:
        # 仅gt切片
        if np.max(gtMat) > 0 :
            try:
                hdcoeff = metric.binary.hd95(predMat, gtMat) * fColSpacing
                hd_list.append(hdcoeff)
            except:
                pass
    return round(np.mean(hd_list), 2) if len(hd_list) > 0 else 0


# ==========================================================
# -------------------- 主评估函数 ---------------------------
# ==========================================================

def evaluate_model_combined(gt_dir, pred_dir, sheet_name, output_excel):
    print(f"\n📊 正在评估模型: {sheet_name}")
    records = []
    cases = sorted([f for f in os.listdir(pred_dir) if f.endswith(".nii.gz")])

    for idx, case_file in enumerate(cases):
        case_id = f"p_{idx}"
        pred_path = os.path.join(pred_dir, case_file)
        gt_path = os.path.join(gt_dir, case_file)
        if not os.path.exists(gt_path):
            print(f"❌ GT 文件不存在：{gt_path}")
            records.append([case_id] + [""] * 16)
            continue

        # ---- 2D ----
        dice2d = ComputeDSC2D(gt_path, pred_path)
        hd95_2d = ComputeHD95(gt_path, pred_path)

        try:
            # ---- 3D Dice / SE / SP ----
            cmd_metrics = f'plastimatch dice "{gt_path}" "{pred_path}"'
            result_metrics = subprocess.run(
                cmd_metrics, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                text=True, shell=True, check=True
            )
            output_metrics = result_metrics.stdout

            def extract(pattern):
                m = re.search(pattern, output_metrics)
                return float(m.group(1)) if m else 0.0

            dice3d = round(extract(r"DICE:\s*([0-9.]+)"), 2)
            se = round(extract(r"SE:\s*([0-9.]+)") * 100, 2)
            sp = round(extract(r"SP:\s*([0-9.]+)") * 100, 2)
            tp = int(extract(r"TP:\s*([0-9.]+)"))
            tn = int(extract(r"TN:\s*([0-9.]+)"))
            fp = int(extract(r"FP:\s*([0-9.]+)"))
            fn = int(extract(r"FN:\s*([0-9.]+)"))

            iou = round(tp / (tp + fp + fn), 2) if (tp + fp + fn) > 0 else 0.0
            precision = round((tp / (tp + fp)) * 100, 2) if (tp + fp) > 0 else 0.0
            accuracy = round(((tp + tn) / (tp + tn + fp + fn)) * 100, 2) if (tp + tn + fp + fn) > 0 else 0.0

            # Step 2: plastimatch Hausdorff boundary 版
            cmd_hd = f'plastimatch dice --hausdorff "{gt_path}" "{pred_path}"'
            result_hd = subprocess.run(cmd_hd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, shell=True,
                                       check=True)
            output_hd = result_hd.stdout
            hd95_match = re.search(r"Percent\s*\(0\.95\)\s*Hausdorff\s*distance\s*\(boundary\)\s*=\s*([0-9.]+)",
                                   output_hd)
            hd95_3d = round(float(hd95_match.group(1)), 2) if hd95_match else ""

            # ---- ASD ----
            gt_img = sitk.ReadImage(gt_path)
            pred_img = sitk.ReadImage(pred_path)
            gt_array = (sitk.GetArrayFromImage(gt_img) > 0).astype(np.uint8)
            pred_array = (sitk.GetArrayFromImage(pred_img) > 0).astype(np.uint8)
            spacing = gt_img.GetSpacing()
            try:
                asd_voxel = metric.binary.asd(pred_array, gt_array)
                asd_val = round(asd_voxel * np.mean(spacing), 2)
            except Exception:
                asd_val = 0.0

            # ---- 体积 ----
            voxel_volume = spacing[0] * spacing[1] * spacing[2]
            ref_vol = round((gt_array > 0).sum() * voxel_volume / 1000, 2)
            pred_vol = round((pred_array > 0).sum() * voxel_volume / 1000, 2)

            records.append([
                case_id, dice2d, hd95_2d,
                dice3d, hd95_3d, iou, asd_val,
                se, sp, precision, accuracy,
                tp, tn, fp, fn, ref_vol, pred_vol
            ])
        except Exception as e:
            print(f"❌ Error: {e}")
            records.append([case_id] + [""] * 16)

    # ---- 表头 ----
    columns = [
        "ID", "2D Dice", "2D HD95 (mm)",
        "3D Dice", "3D HD95 (mm)", "IoU", "ASD (mm)",
        "Sensitivity (%)", "Specificity (%)", "Precision (%)", "Accuracy (%)",
        "TP (vox)", "TN (vox)", "FP (vox)", "FN (vox)",
        "Ref_volume (cc)", "Pred_volume (cc)"
    ]

    df = pd.DataFrame(records, columns=columns)

    # ---- 去除旧的 Mean / STD 行，仅保留 p_数字行 ----
    df = df[~df["ID"].isin(["Mean", "STD"])]

    # ✅ 彻底断开引用，防止 Mean 行被隐式包含
    numeric_df = (
        df[df["ID"].str.match(r"p_\d+$", na=False)]
        .reset_index(drop=True)
        .copy()
    )

    if numeric_df.empty:
        print("⚠️ 警告：未匹配到任何 p_xx 行，STD/Mean 将为空！")

    metric_cols = ["2D DSC", "2D HD95 (mm)", "3D Dice", "3D HD95 (mm)", "IoU", "ASD (mm)"]

    # ---- 计算 Mean & STD ----
    mean_row, std_row = ["Mean"], ["STD"]

    # ✅ 自动筛选出 p_ 开头的病例行（不包含 Mean / STD）
    numeric_df = (
        df[df["ID"].astype(str).str.match(r"^p_\d+$", na=False)]
        .reset_index(drop=True)
        .copy()
    )

    # 防止空 DataFrame
    if numeric_df.empty:
        print(f"⚠️ 警告：未找到 p_ 开头的病例行，跳过统计计算。")
    else:
        for i, col in enumerate(columns[1:], start=1):
            vals = pd.to_numeric(numeric_df[col], errors='coerce')
            mean_val = np.nanmean(vals)
            # ✅ 使用样本标准差，与 Excel STDEV 一致
            std_val = np.nanstd(vals, ddof=1)

            # 仅计算前6列的标准差
            if 1 <= i <= 6:
                mean = round(mean_val, 2) if not np.isnan(mean_val) else ""
                std = round(std_val, 2) if not np.isnan(std_val) else ""
            else:
                mean = round(mean_val, 2) if not np.isnan(mean_val) else ""
                std = ""  # 后面列空出来

            mean_row.append(mean)
            std_row.append(std)

        # 添加均值和标准差行
        df.loc[len(df)] = mean_row
        df.loc[len(df)] = std_row

    # ---- 写入 Excel ----
    mode = 'a' if os.path.exists(output_excel) else 'w'
    with pd.ExcelWriter(output_excel, engine='openpyxl', mode=mode) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    # ---- 样式 ----
    wb = openpyxl.load_workbook(output_excel)
    ws = wb[sheet_name]
    for i, col in enumerate(columns, start=1):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[get_column_letter(i)])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(output_excel)

    # ---- 控制台摘要 ----
    print(f"[{sheet_name}] 2D Dice={mean_row[1]}±{std_row[1]} | 2D HD95={mean_row[2]}±{std_row[2]} | "
          f"3D Dice={mean_row[3]}±{std_row[3]} | 3D HD95={mean_row[4]}±{std_row[4]} | ASD={mean_row[6]}±{std_row[6]}")


# ==========================================================
# -------------------- Summary 汇总 ------------------------
# ==========================================================

def generate_summary_mean_sheet(output_excel):
    wb = openpyxl.load_workbook(output_excel)
    summary_data, ref_columns = [], []

    for sheet in wb.sheetnames:
        if sheet == "Summary_Mean":
            continue
        ws = wb[sheet]
        if ws.max_row < 3:
            continue
        if not ref_columns:
            ref_columns = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        mean_row = [ws.cell(row=ws.max_row - 1, column=i).value for i in range(1, ws.max_column + 1)]
        std_row = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]
        model = sheet
        combined = [model]
        metric_cols = ["2D DSC", "2D HD95 (mm)", "3D Dice", "3D HD95 (mm)", "IoU", "ASD (mm)"]
        for i, col in enumerate(ref_columns[1:], start=1):
            mean_val, std_val = mean_row[i], std_row[i]
            if col in metric_cols:
                combined.append(f"{mean_val:.2f} ± {std_val:.2f}")
            else:
                combined.append(f"{mean_val:.2f}" if isinstance(mean_val, (int, float, np.float64)) else "")
        summary_data.append(combined)

    columns = ["模型名称"] + ref_columns[1:]
    df = pd.DataFrame(summary_data, columns=columns)

    if "Summary_Mean" in wb.sheetnames:
        wb.remove(wb["Summary_Mean"])
    ws = wb.create_sheet("Summary_Mean", 0)

    # ---- 写入 DataFrame ----
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # ---- 表头加粗（必须放在写完 DataFrame 之后）----
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- 自动列宽 ----
    for col_idx, col in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_excel)
    print("✅ Summary_Mean 生成完成！")


# ==========================================================
# -------------------- 主入口 -------------------------------
# ==========================================================

if __name__ == "__main__":
    gtDir = r"D:\SAM\GTVp_CTonly\20250809\nnUNet\Dataset001_GTVp\labelsTs"

    # cm
    output_excel = r"C:\Users\WS\Desktop\cm_eval_all.xlsx"
    base_root = r"C:\Users\WS\Desktop\20251104\cm"
    base_dict = {
        # "Freeze_encoder_decoder": "encoder_decoder",
        # "Freeze_image_encoder": "image_encoder",
        # "Freeze_mask_decoder": "mask_decoder",
        "TrainAll": "TrainAll"
    }
    x_lists = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.2, 1.5]
    model_paths = {
        f"{short}_{x}": os.path.join(
            base_root,
            base,
            f"expand_{int(x)}cm" if x == 0 else f"expand_{x:.1f}cm"
        )
        for base, short in base_dict.items()
        for x in x_lists
    }

    # # Num
    # output_excel = r"C:\Users\WS\Desktop\Num_eval.xlsx"
    # model_paths = {
    #     "2slices": r"C:\Users\WS\Desktop\20251104\Num_box_prompts\2_slices",
    #     "3slices": r"C:\Users\WS\Desktop\20251104\Num_box_prompts\3_slices",
    #     "5slices": r"C:\Users\WS\Desktop\20251104\Num_box_prompts\5_slices",
    #     "7slices": r"C:\Users\WS\Desktop\20251104\Num_box_prompts\7_slices",
    #     "all_slices": r"C:\Users\WS\Desktop\20251104\cm\Freeze_image_encoder\expand_0.5cm",
    # }

    # # Pos
    # output_excel = r"C:\Users\WS\Desktop\Pos_eval_random.xlsx"
    # model_paths = {
    #     # "middle_z": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\middle_z_rep0",
    #     # "max_area": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\max_area_rep0",
    #     # "middle_volume": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\mid_volume_rep0",
    #     "25": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep25",
    #     "104": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep104",
    #     "114": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep114",
    #     "142": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep142",
    #     "228": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep228",
    #     "250": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep250",
    #     "281": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep281",
    #     "654": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep654",
    #     "754": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep754",
    #     "759": r"C:\Users\WS\Desktop\20251104\Pos_box_prompts\random_rep759",
    # }

    # # Baseline
    # output_excel = r"C:\Users\WS\Desktop\Baseline_eval2.xlsx"
    # model_paths = {
    #     # "nnUNet_2d": r"C:\Users\WS\Desktop\20251104\baseline\nnUNet_2d",
    #     # "nnUNet_3d": r"C:\Users\WS\Desktop\20251104\baseline\nnUNet_3d",
    #     "UNet": r"C:\Users\WS\Desktop\20251104\baseline\UNet",
    #     # "Deeplabv3+": r"C:\Users\WS\Desktop\20251104\baseline\Deeplabv3+",
    #     # "TransUNet": r"C:\Users\WS\Desktop\20251104\baseline\TransUNet",
    #     # "SwinUNet": r"C:\Users\WS\Desktop\20251104\baseline\SwinUNet",
    # }

    for name, path in model_paths.items():
        evaluate_model_combined(gtDir, path, name, output_excel)

    generate_summary_mean_sheet(output_excel)
    print(f"\n🎯 所有模型评估完成！结果保存在：{output_excel}")
