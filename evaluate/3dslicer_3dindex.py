import os
import subprocess
import re
import SimpleITK as sitk
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def evaluate_prediction(gt_dir, pred_dir, sheet_name, output_excel):
    records = []
    cases = sorted([f for f in os.listdir(pred_dir) if f.endswith(".nii.gz")])

    for idx, case_file in enumerate(cases):
        case_id = f"p_{idx}"
        pred_path = os.path.join(pred_dir, case_file)
        gt_path = os.path.join(gt_dir, case_file)

        if not os.path.exists(gt_path):
            print(f"\u274c GT 文件不存在：{gt_path}")
            records.append([case_id] + [""] * 13)
            continue

        try:
            cmd_metrics = f'plastimatch dice "{gt_path}" "{pred_path}"'
            result_metrics = subprocess.run(cmd_metrics, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, shell=True, check=True)
            output_metrics = result_metrics.stdout

            def extract(pattern):
                match = re.search(pattern, output_metrics)
                return float(match.group(1)) if match else 0.0

            dice = round(extract(r"DICE:\s*([0-9.]+)"), 2)
            se = round(extract(r"SE:\s*([0-9.]+)") * 100, 2)
            sp = round(extract(r"SP:\s*([0-9.]+)") * 100, 2)
            tp = int(extract(r"TP:\s*([0-9.]+)"))
            tn = int(extract(r"TN:\s*([0-9.]+)"))
            fp = int(extract(r"FP:\s*([0-9.]+)"))
            fn = int(extract(r"FN:\s*([0-9.]+)"))

            iou = round(tp / (tp + fp + fn), 2) if (tp + fp + fn) > 0 else 0.0
            precision = round((tp / (tp + fp)) * 100, 2) if (tp + fp) > 0 else 0.0
            accuracy = round(((tp + tn) / (tp + tn + fp + fn)) * 100, 2) if (tp + tn + fp + fn) > 0 else 0.0

            cmd_hd = f'plastimatch dice --hausdorff "{gt_path}" "{pred_path}"'
            result_hd = subprocess.run(cmd_hd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, shell=True, check=True)
            output_hd = result_hd.stdout
            hd95_match = re.search(r"Percent \(0.95\) Hausdorff distance \(boundary\)\s*=\s*([0-9.]+)", output_hd)
            hd95 = round(float(hd95_match.group(1)), 2) if hd95_match else ""

            gt_img = sitk.ReadImage(gt_path)
            pred_img = sitk.ReadImage(pred_path)
            spacing = gt_img.GetSpacing()
            voxel_volume = spacing[0] * spacing[1] * spacing[2]
            gt_array = sitk.GetArrayFromImage(gt_img)
            pred_array = sitk.GetArrayFromImage(pred_img)
            ref_vol = round((gt_array > 0).sum() * voxel_volume / 1000, 2)
            pred_vol = round((pred_array > 0).sum() * voxel_volume / 1000, 2)

            records.append([
                case_id, dice, hd95, iou,
                se, sp, precision, accuracy,
                tp, tn, fp, fn,
                ref_vol, pred_vol
            ])

        except Exception as e:
            print(f"\u274c Error processing {case_file}: {e}")
            records.append([case_id] + [""] * 13)

    columns = [
        "ID", "Dice", "HD95 (mm)", "IoU",
        "Sensitivity (%)", "Specificity (%)",
        "Precision (%)", "Accuracy (%)",
        "TP (vox)", "TN (vox)", "FP (vox)", "FN (vox)",
        "Ref_volume (cc)", "Pred_volume (cc)"
    ]
    df = pd.DataFrame(records, columns=columns)

    mean_row = ["Mean"] + [round(df[col].astype(float).mean(), 2) if col != "ID" else "" for col in columns[1:]]
    df.loc[len(df)] = mean_row

    file_exists = os.path.exists(output_excel)
    mode = 'a' if file_exists else 'w'
    engine = 'openpyxl' if file_exists else 'xlsxwriter'

    writer_args = {
        'path': output_excel,
        'engine': engine,
        'mode': mode
    }
    if file_exists:
        writer_args['if_sheet_exists'] = 'replace'

    with pd.ExcelWriter(**writer_args) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    # 手动打开 workbook 并设置格式
    wb = load_workbook(output_excel)
    ws = wb[sheet_name]

    for i, col in enumerate(columns):
        col_letter = get_column_letter(i + 1)
        col_width = min(max(len(col) + 2, 10), 25)
        ws.column_dimensions[col_letter].width = col_width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_excel)

# summary生成
def generate_summary_mean_sheet(output_excel):
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    import pandas as pd

    wb = openpyxl.load_workbook(output_excel)
    summary_data = []

    ref_columns = []
    for sheet in wb.sheetnames:
        if sheet == "Summary_Mean":
            continue
        ws = wb[sheet]
        if ws.max_row < 2:
            continue
        if not ref_columns:
            ref_columns = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

        last_row = [ws.cell(row=ws.max_row, column=col).value for col in range(1, ws.max_column + 1)]
        model_name = sheet
        expand_pixel = ""
        for item in ["0p", "3p", "5p", "7p", "9p"]:
            if item in sheet:
                expand_pixel = item.replace("p", "")
                break
        clean_model_name = model_name.rsplit('_', 1)[0] if model_name.startswith('SAM_') else model_name
        new_row = [clean_model_name, expand_pixel] + last_row[1:]
        summary_data.append(new_row)

    columns = ["模型名称", "框外扩（pixel）"] + ref_columns[1:]
    summary_df = pd.DataFrame(summary_data, columns=columns)

    # 删除已存在的 Summary_Mean sheet（如果存在）
    if "Summary_Mean" in wb.sheetnames:
        std = wb["Summary_Mean"]
        wb.remove(std)
    new_sheet = wb.create_sheet(title="Summary_Mean", index=0)

    # 设置列宽按最宽字符计算，标题加粗居中
    col_max_widths = [len(str(col)) for col in summary_df.columns]
    for row in summary_df.values:
        for idx, val in enumerate(row):
            val_len = len(str(val)) if val is not None else 0
            if val_len > col_max_widths[idx]:
                col_max_widths[idx] = val_len

    for i, col in enumerate(summary_df.columns):
        col_letter = get_column_letter(i+1)
        width = min(col_max_widths[i] + 2, 35)
        new_sheet.column_dimensions[col_letter].width = width
        cell = new_sheet.cell(row=1, column=i+1, value=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    # 写入数据并居中
    for row_idx in range(2, len(summary_df) + 2):
        for col_idx in range(1, len(summary_df.columns) + 1):
            cell = new_sheet.cell(row=row_idx, column=col_idx, value=summary_df.iat[row_idx - 2, col_idx - 1])
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 合并模型名称相同的单元格
    summary_ws = wb["Summary_Mean"]
    current_model = None
    start_row = 2
    for row in range(2, summary_ws.max_row + 2):
        model_cell = summary_ws.cell(row=row, column=1)
        model_value = model_cell.value
        if model_value != current_model:
            if current_model is not None and row - start_row > 1:
                summary_ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                summary_ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            current_model = model_value
            start_row = row
    if current_model is not None and summary_ws.max_row - start_row >= 1:
        summary_ws.merge_cells(start_row=start_row, start_column=1, end_row=summary_ws.max_row, end_column=1)
        summary_ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_excel)



# ====== 调用入口 ======
if __name__ == '__main__':
    gt_dir = r"C:\Users\dell\Desktop\SAM\GTVp_CTonly\20250809\nnUNet\Dataset001_GTVp\labelsTs"
    # gt_dir = r"C:\\Users\\WS\\Desktop\\20250809\\nnUNet\\Dataset002_RGB\\labelsTs"
    output_excel = r"C:\Users\dell\Desktop/three_prompt/expand_0.5cm_top_bottom_vis/eval_3d_top_bottom.xlsx"

    # # nnunet评估输出
    # pred_dir1 = r"C:\\Users\\dell\\Desktop\\CTV\\testresults\\nnUNet_CTV_2d"
    # evaluate_prediction(gt_dir, pred_dir1, "nnUNet_2d", output_excel)

    # pred_dir2 = r"C:\\Users\\dell\\Desktop\\CTV\\testresults\\nnUNet_CTV_3d"
    # evaluate_prediction(gt_dir, pred_dir2, "nnUNet_3d", output_excel)

    # # SAM评估
    # base_name = "TrainAll"
    # base_dir = fr"C:\Users\dell\Desktop\cm\{base_name}"
    # expand_cm_list = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.2, 1.5]
    #
    # for cm in expand_cm_list:
    #     cm_str = str(cm)
    #     pred_dir = os.path.join(base_dir, f"expand_{cm}cm")  # 文件夹是真实 cm 名字
    #     sheet_name = f"SAM_{base_name}_{cm_str}cm"
    #     evaluate_prediction(gt_dir, pred_dir, sheet_name, output_excel)

    # SAM 稀疏提示评估输出
    pred_dir1 = r"C:\Users\dell\Desktop\three_prompt/expand_0.5cm_top_bottom_vis"
    evaluate_prediction(gt_dir, pred_dir1, "SAM", output_excel)

    # 汇总平均值
    generate_summary_mean_sheet(output_excel)

