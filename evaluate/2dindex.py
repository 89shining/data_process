"""
计算GT和pred均>0的切片
"""


import numpy as np
import os
import SimpleITK as sitk
from medpy import metric
import pandas as pd

def DiceCoefficient(a, b):
    """dice coefficient 2nt/na + nb."""
    SMOOTH = 1e-5
    a = a > 0.5
    b = b > 0.5
    intersection = np.sum(np.logical_and(a, b))
    return (2. * intersection + SMOOTH) / (np.sum(a) + np.sum(b) + SMOOTH)

def ComputeDSC2D(gtDir, predDir):
    gtVolume = sitk.GetArrayFromImage(sitk.ReadImage(gtDir))
    predVolume = sitk.GetArrayFromImage(sitk.ReadImage(predDir))
    dice = 0.0
    iNum = 0
    for i in range(predVolume.shape[0]):
        gtMat = gtVolume[i]
        predMat = predVolume[i]

        if np.max(gtMat) == 0 or np.max(predMat) == 0:
            continue  # 跳过空结构的切片

        gtMat[gtMat > 0] = 1
        predMat[predMat > 0] = 1

        dice += DiceCoefficient(predMat, gtMat)
        iNum += 1
    if iNum == 0:
        return 0
    dice = dice / iNum
    return dice

def ComputeHD95(gtDir, predDir):
    oridata = sitk.ReadImage(gtDir)
    fColSpacing = oridata.GetSpacing()[1]
    gtVolume = sitk.GetArrayFromImage(oridata)
    predVolume = sitk.GetArrayFromImage(sitk.ReadImage(predDir))

    hdResult = []
    for i in range(predVolume.shape[0]):
        predMat = predVolume[i]
        predMat[np.where(predMat > 0)] = 1
        gtMat = gtVolume[i]
        gtMat[np.where(gtMat > 0)] = 1

        if np.max(gtMat) > 0 and np.max(predMat) > 0:
            try:
                hdcoeff = metric.binary.hd95(predMat, gtMat) * fColSpacing
                hdResult.append(hdcoeff)
            except:
                pass

    result = np.mean(hdResult)
    return result

def evaluate_model_to_excel(gtDir, predDir, model_name, excel_path):
    """
       将某个模型的 2D Dice 和 HD95 评估写入 Excel 的一个 sheet
       """
    print(f"\n📊 正在评估模型: {model_name}")
    ids, dices, hd95s = [], [], []

    palist = [f for f in os.listdir(predDir) if f.endswith('.nii.gz')]
    for pa in palist:
        pred_path = os.path.join(predDir, pa)
        gt_path = os.path.join(gtDir, pa)

        dice = ComputeDSC2D(gt_path, pred_path)
        hd95 = ComputeHD95(gt_path, pred_path)

        patient_num = int(pa.split('_')[-1].split('.')[0])
        patient_id = f"p_{patient_num}"

        ids.append(patient_id)
        dices.append(dice)
        hd95s.append(hd95)

        print(f"  {patient_id}: Dice={np.round(dice, 2)}, HD95={np.round(hd95, 2)}")

    # 添加平均值
    ids.append("Mean")
    dices.append(np.mean(dices))
    hd95s.append(np.mean(hd95s))

    df = pd.DataFrame({
        'ID': ids,
        '2D DSC': np.round(dices, 2),
        '2D HD95': np.round(hd95s, 2)
    })

    # 写入 Excel，追加 sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a' if os.path.exists(excel_path) else 'w') as writer:
        df.to_excel(writer, index=False, sheet_name=model_name)

    print(f"写入完成 → Sheet: {model_name}")

# summary生成
def generate_summary_mean_sheet(output_excel):
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    import pandas as pd

    wb = openpyxl.load_workbook(output_excel)
    summary_data = []

    ref_columns = []
    for sheet in wb.sheetnames:
        if sheet == "Summary_Mean":
            continue
        ws = wb[sheet]
        if ws.max_row < 2:
            continue
        if not ref_columns:
            ref_columns = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

        last_row = [ws.cell(row=ws.max_row, column=col).value for col in range(1, ws.max_column + 1)]
        model_name = sheet
        expand_pixel = ""
        for item in ["0p", "3p", "5p", "7p", "9p"]:
            if item in sheet:
                expand_pixel = item.replace("p", "")
                break
        clean_model_name = model_name.rsplit('_', 1)[0] if model_name.startswith('SAM_') else model_name
        new_row = [clean_model_name, expand_pixel] + last_row[1:]
        summary_data.append(new_row)

    columns = ["模型名称", "框外扩（pixel）"] + ref_columns[1:]
    summary_df = pd.DataFrame(summary_data, columns=columns)

    # 删除已存在的 Summary_Mean sheet（如果存在）
    if "Summary_Mean" in wb.sheetnames:
        std = wb["Summary_Mean"]
        wb.remove(std)
    new_sheet = wb.create_sheet(title="Summary_Mean", index=0)

    # 设置列宽按最宽字符计算，标题加粗居中
    col_max_widths = [len(str(col)) for col in summary_df.columns]
    for row in summary_df.values:
        for idx, val in enumerate(row):
            val_len = len(str(val)) if val is not None else 0
            if val_len > col_max_widths[idx]:
                col_max_widths[idx] = val_len

    for i, col in enumerate(summary_df.columns):
        col_letter = get_column_letter(i+1)
        width = min(col_max_widths[i] + 2, 35)
        new_sheet.column_dimensions[col_letter].width = width
        cell = new_sheet.cell(row=1, column=i+1, value=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    # 写入数据并居中
    for row_idx in range(2, len(summary_df) + 2):
        for col_idx in range(1, len(summary_df.columns) + 1):
            cell = new_sheet.cell(row=row_idx, column=col_idx, value=summary_df.iat[row_idx - 2, col_idx - 1])
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 合并模型名称相同的单元格
    summary_ws = wb["Summary_Mean"]
    current_model = None
    start_row = 2
    for row in range(2, summary_ws.max_row + 2):
        model_cell = summary_ws.cell(row=row, column=1)
        model_value = model_cell.value
        if model_value != current_model:
            if current_model is not None and row - start_row > 1:
                summary_ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                summary_ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            current_model = model_value
            start_row = row
    if current_model is not None and summary_ws.max_row - start_row >= 1:
        summary_ws.merge_cells(start_row=start_row, start_column=1, end_row=summary_ws.max_row, end_column=1)
        summary_ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_excel)


if __name__ == "__main__":
    # GT文件目录
    gtDir = r'C:\Users\dell\Desktop\SAM\GTVp_CTonly\20250809\nnUNet\Dataset001_GTVp\labelsTs'
    # excel文件保存地址
    excel_path = r"C:\Users\dell\Desktop/three_prompt/expand_0.5cm_top_bottom_vis/eval_2d_top_bottom.xlsx"

    # 模型及其预测路径配置
    base_name = "TrainAll"
    # 外扩pixel
    model_paths = {
        # "nnUNet_2d": r'C:\Users\dell\Desktop\CTV\testresults\nnUNet_CTV_2d',
        # # "nnUNet_2.5d": r'C:\Users\WS\Desktop\20250809\testresults\nnUNet_2.5d'
        # "nnUNet_3d": r'C:\Users\dell\Desktop\CTV\testresults\nnUNet_CTV_3d',
        # f"SAM_{base_name}_0p": fr'C:\Users\dell\Desktop\cm\testresults\{base_name}/expand_0p',
        # f"SAM_{base_name}_3p": fr'C:\Users\dell\Desktop\CTV\testresults\{base_name}/expand_3p',
        # f"SAM_{base_name}_5p": fr'C:\Users\dell\Desktop\CTV\testresults\{base_name}/expand_5p',
        # f"SAM_{base_name}_7p": fr'C:\Users\dell\Desktop\CTV\testresults\{base_name}/expand_7p',
        # f"SAM_{base_name}_9p": fr'C:\Users\dell\Desktop\CTV\testresults\{base_name}/expand_9p',
        "SAM": r"C:\Users\dell\Desktop\three_prompt/expand_0.5cm_top_bottom_vis"
    }

    # # 外扩 cm 设置
    # base_cm_list = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.2, 1.5]
    # model_paths = {}
    # for cm in base_cm_list:
    #     cm_str = str(cm)
    #     model_name = f"SAM_{base_name}_{cm_str}cm"
    #     predDir = fr'C:\Users\dell\Desktop\cm\{base_name}\expand_{cm}cm'
    #     model_paths[model_name] = predDir

    # 依次写入每个模型评估结果
    for model_name, predDir in model_paths.items():
        evaluate_model_to_excel(gtDir, predDir, model_name, excel_path)

    print(f"\n 所有模型评估完成，结果保存在：{excel_path}")

    # 汇总平均值
    generate_summary_mean_sheet(excel_path)


