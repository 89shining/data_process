# -*- coding: utf-8 -*-
"""
合并多次随机实验的患者级指标均值，并在底部计算20位患者的Mean/STD（ddof=1）
- 忽略第一个sheet（总表）
- 仅保留ID形如 p_0~p_19 的行，过滤Mean/STD等汇总行
- 所有数值保留两位小数，Excel中设置为0.00，单元格居中
"""

import re
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Optional, List
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def _round_numeric_to_2(df: pd.DataFrame) -> pd.DataFrame:
    """仅对数值列四舍五入到两位小数"""
    out = df.copy()
    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            out[c] = np.round(out[c].astype(float), 2)
    return out


def _write_center_and_2dec(ws):
    """将整张表数据居中，数值列设置0.00"""
    max_row, max_col = ws.max_row, ws.max_column
    # 先整体居中
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    # 数值列统一0.00：除第一列(ID)外
    for col_idx in range(2, max_col + 1):
        col_letter = get_column_letter(col_idx)
        for r in range(2, max_row + 1):  # 跳过表头
            ws[f"{col_letter}{r}"].number_format = "0.00"


def per_patient_mean_table(
    input_xlsx: str,
    output_xlsx: Optional[str] = None,           # 不给就自动生成 *_per_patient.xlsx
    sheet_names: Optional[List[str]] = None,     # 不给就用除第一个外的全部sheet
    out_sheet_name: str = "PerPatient_Mean",
) -> str:
    in_path = Path(input_xlsx)
    if not in_path.exists():
        raise FileNotFoundError(f"找不到输入文件：{in_path}")

    # 自动输出路径
    out_path = Path(output_xlsx) if output_xlsx else in_path.with_name(f"{in_path.stem}_per_patient.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = out_path.exists()

    # 读取所有sheet
    all_sheets: dict = pd.read_excel(in_path, sheet_name=None, engine="openpyxl")
    sheet_list = list(all_sheets.keys())
    sheets_to_use = sheet_list[1:] if sheet_names is None else sheet_names
    if not sheets_to_use:
        raise ValueError("没有可用的sheet（除第一个总表外）。")

    # 逐sheet收集患者行
    id_pattern = re.compile(r"^p_\d+$", re.IGNORECASE)
    frames = []
    id_col_name = None

    for s in sheets_to_use:
        df = all_sheets[s].copy()

        # 找到ID列（不区分大小写）
        id_candidates = [c for c in df.columns if str(c).strip().lower() == "id"]
        if not id_candidates:
            raise ValueError(f"Sheet '{s}' 未找到 'ID' 列。")
        id_col = id_candidates[0]
        id_col_name = id_col  # 记住原始列名

        # 过滤仅保留 p_*
        df = df[df[id_col].astype(str).str.match(id_pattern, na=False)]

        # 设为索引，转数值列
        df = df.set_index(id_col)
        for c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        frames.append(df)

    # 合并并按患者ID在列上求均值（跨实验）
    cat = pd.concat(frames, axis=0, sort=False)
    patient_means = cat.groupby(cat.index).mean(numeric_only=True)
    # ===== 强制按患者编号数值排序（p_0, p_1, ..., p_39）=====
    patient_means = (
        patient_means
        .assign(_pid_num=patient_means.index.str.extract(r"(\d+)").astype(int).values)
        .sort_values("_pid_num")
        .drop(columns="_pid_num")
    )

    # 生成底部的 Mean / STD（对20位患者按列统计；/n-1）
    bottom_mean = patient_means.mean(axis=0, numeric_only=True)
    bottom_std  = patient_means.std(axis=0, ddof=1, numeric_only=True)

    # 四舍五入到两位
    patient_means = _round_numeric_to_2(patient_means)
    bottom_mean   = np.round(bottom_mean.astype(float), 2)
    bottom_std    = np.round(bottom_std.astype(float), 2)

    # 组装输出表（加上ID列 + 底部Mean/STD行）
    out_df = patient_means.reset_index()
    out_df.rename(columns={"index": id_col_name or "ID"}, inplace=True)
    # 追加Mean/STD两行
    mean_row = {col: (bottom_mean[col] if col in bottom_mean.index else None) for col in out_df.columns}
    std_row  = {col: (bottom_std[col]  if col in bottom_std.index  else None) for col in out_df.columns}
    mean_row[out_df.columns[0]] = "Mean"
    std_row[out_df.columns[0]]  = "STD"
    out_df = pd.concat([out_df, pd.DataFrame([mean_row, std_row])], ignore_index=True)

    # 写入Excel：若文件存在则替换同名sheet，否则新建
    writer_kwargs = dict(
        engine="openpyxl",
        mode=("a" if file_exists else "w"),
        if_sheet_exists=("replace" if file_exists else None),
    )
    with pd.ExcelWriter(out_path, **writer_kwargs) as w:
        out_df.to_excel(w, sheet_name=out_sheet_name, index=False)
        ws = w.sheets[out_sheet_name]
        _write_center_and_2dec(ws)

    return str(out_path)


if __name__ == "__main__":
    # ======= 改这里：输入 Excel 路径 =======
    input_file = r"C:\Users\dell\Desktop\20251224_Test40\Results_python\Pos_eval_random10.xlsx"
    # 不填 output_xlsx 会在同目录生成 *_per_patient.xlsx
    saved = per_patient_mean_table(
        input_xlsx=input_file,
        output_xlsx=None,          # 自动生成
        sheet_names=None,          # 用除第一个外的全部sheet
        out_sheet_name="PerPatient_Mean"
    )
    print("已保存到：", saved)
