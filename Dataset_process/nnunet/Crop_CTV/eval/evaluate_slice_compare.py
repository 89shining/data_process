"""
Per-slice comparison between nnUNet_crop_restore and nnUNet_all_volume.

Only slices with GT foreground are evaluated.
For each (case, z) row, output:
- ID
- current z index
- upper bound z index
- lower bound z index (smaller index)
- normalized z position (smaller means closer to upper bound)
- Dice for crop_restore and all_volume + delta (crop_restore - all_volume)
- HD95 for crop_restore and all_volume + delta (all_volume - crop_restore)

Output: one Excel table (single sheet).
"""

import math
import os
import re
from typing import Dict, List, Tuple

import numpy as np
import openpyxl
import SimpleITK as sitk
from medpy import metric
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def strip_nii_ext(name: str) -> str:
    if name.endswith(".nii.gz"):
        return name[:-7]
    if name.endswith(".nii"):
        return name[:-4]
    return name


def extract_numeric_key(name: str) -> str:
    stem = strip_nii_ext(os.path.basename(name))
    nums = re.findall(r"\d+", stem)
    if not nums:
        raise ValueError(f"No numeric id found in filename: {name}")
    return str(int(nums[-1]))


def list_nii_files(folder: str) -> List[str]:
    return sorted([f for f in os.listdir(folder) if f.endswith(".nii.gz") or f.endswith(".nii")])


def build_numeric_index(folder: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for name in list_nii_files(folder):
        key = extract_numeric_key(name)
        path = os.path.join(folder, name)
        if key in out:
            raise ValueError(
                f"Duplicate numeric id in folder {folder}: id={key}, "
                f"files={os.path.basename(out[key])} and {name}"
            )
        out[key] = path
    return out


def dice_2d(pred_slice: np.ndarray, gt_slice: np.ndarray) -> float:
    pred = pred_slice > 0
    gt = gt_slice > 0
    inter = np.logical_and(pred, gt).sum()
    smooth = 1e-5
    return float((2.0 * inter + smooth) / (pred.sum() + gt.sum() + smooth))


def diag_mm_2d(shape_yx: Tuple[int, int], spacing_xy: Tuple[float, float]) -> float:
    y, x = shape_yx
    sx, sy = spacing_xy
    return float(math.sqrt((y * sy) ** 2 + (x * sx) ** 2))


def hd95_2d_mm(pred_slice: np.ndarray, gt_slice: np.ndarray, spacing_xy: Tuple[float, float]) -> float:
    pred = (pred_slice > 0).astype(np.uint8)
    gt = (gt_slice > 0).astype(np.uint8)

    pred_any = bool(pred.max() > 0)
    gt_any = bool(gt.max() > 0)

    if pred_any and gt_any:
        # medpy expects voxelspacing matching array axis order (y, x)
        return float(metric.binary.hd95(pred, gt, voxelspacing=(spacing_xy[1], spacing_xy[0])))

    # Only GT-positive slices are evaluated; if pred is empty, use 2D diagonal penalty.
    return diag_mm_2d(pred.shape, spacing_xy)


def auto_fit_and_center(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[letter].width = min(max_len + 2, 45)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def evaluate_slice_table(
    gt_dir: str,
    all_volume_dir: str,
    crop_restore_dir: str,
    out_excel: str,
) -> None:
    gt_index = build_numeric_index(gt_dir)
    all_index = build_numeric_index(all_volume_dir)
    crop_index = build_numeric_index(crop_restore_dir)

    columns = [
        "ID",
        "Current Z",
        "Upper Z",
        "Lower Z",
        "Norm Z (0=upper, 1=lower)",
        "Dice all_volume",
        "Dice crop_restore",
        "Dice Delta (crop_restore - all_volume)",
        "HD95 all_volume (mm)",
        "HD95 crop_restore (mm)",
        "HD95 Delta (all_volume - crop_restore)",
    ]

    rows: List[List[object]] = []

    common_keys = sorted(set(gt_index.keys()) & set(all_index.keys()) & set(crop_index.keys()), key=lambda x: int(x))
    missing_all = sorted(set(gt_index.keys()) - set(all_index.keys()), key=lambda x: int(x))
    missing_crop = sorted(set(gt_index.keys()) - set(crop_index.keys()), key=lambda x: int(x))

    if missing_all:
        print(f"[Warn] Missing in all_volume: {missing_all}")
    if missing_crop:
        print(f"[Warn] Missing in crop_restore: {missing_crop}")

    for key in common_keys:
        case_id = f"p_{int(key)}"

        gt_img = sitk.ReadImage(gt_index[key])
        all_img = sitk.ReadImage(all_index[key])
        crop_img = sitk.ReadImage(crop_index[key])

        gt = sitk.GetArrayFromImage(gt_img)
        all_pred = sitk.GetArrayFromImage(all_img)
        crop_pred = sitk.GetArrayFromImage(crop_img)

        if gt.shape != all_pred.shape or gt.shape != crop_pred.shape:
            print(
                f"[Skip] Shape mismatch id={key}: gt={gt.shape}, all={all_pred.shape}, crop={crop_pred.shape}"
            )
            continue

        spacing_xyz = gt_img.GetSpacing()  # (sx, sy, sz)
        spacing_xy = (float(spacing_xyz[0]), float(spacing_xyz[1]))

        gt_pos_z = [z for z in range(gt.shape[0]) if np.any(gt[z] > 0)]
        if not gt_pos_z:
            continue

        lower_z = int(min(gt_pos_z))
        upper_z = int(max(gt_pos_z))
        denom = upper_z - lower_z

        # Sort from upper to lower so norm z increases downward.
        for z in sorted(gt_pos_z, reverse=True):
            gt_slice = gt[z]
            all_slice = all_pred[z]
            crop_slice = crop_pred[z]

            if denom == 0:
                norm_z = 0.0
            else:
                norm_z = float((upper_z - z) / denom)

            dice_all = dice_2d(all_slice, gt_slice)
            dice_crop = dice_2d(crop_slice, gt_slice)
            dice_delta = dice_crop - dice_all

            hd95_all = hd95_2d_mm(all_slice, gt_slice, spacing_xy)
            hd95_crop = hd95_2d_mm(crop_slice, gt_slice, spacing_xy)
            hd95_delta = hd95_all - hd95_crop

            rows.append(
                [
                    case_id,
                    int(z),
                    upper_z,
                    lower_z,
                    norm_z,
                    round(dice_all, 2),
                    round(dice_crop, 2),
                    round(dice_delta, 2),
                    round(hd95_all, 2),
                    round(hd95_crop, 2),
                    round(hd95_delta, 2),
                ]
            )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Slice_Compare"

    ws.append(columns)
    for r in rows:
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Numeric formatting
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = "0.0000"  # norm z
        for c in range(6, 12):
            ws.cell(row=r, column=c).number_format = "0.00"

    auto_fit_and_center(ws)
    wb.save(out_excel)

    print(f"Saved: {out_excel}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    base_dir = r"C:\Users\dell\Desktop\Rectal_146"
    gt_dir = os.path.join(base_dir, "labelsTs")
    all_volume_dir = os.path.join(base_dir, "nnUNet_all_volume")
    crop_restore_dir = os.path.join(base_dir, "nnUNet_crop_restore")

    out_excel = os.path.join(base_dir, "GTslice_Compare_.xlsx")

    evaluate_slice_table(
        gt_dir=gt_dir,
        all_volume_dir=all_volume_dir,
        crop_restore_dir=crop_restore_dir,
        out_excel=out_excel,
    )
