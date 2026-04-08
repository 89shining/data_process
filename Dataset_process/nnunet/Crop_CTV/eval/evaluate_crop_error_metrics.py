"""
Evaluate crop-error fullsize predictions and save Excel tables.

Design goals:
1) Metric definitions and Excel style follow:
   Dataset_process/nnunet/Crop_CTV/eval/evaluate_metrics.py
2) Data matching is adapted for crop-error folders:
   one sheet per error folder (e.g., K1_inward), summary in first sheet.
3) CUDA acceleration is supported when CuPy is available.
"""

import argparse
import math
import os
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import SimpleITK as sitk
from medpy import metric
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from scipy.ndimage import binary_erosion as np_binary_erosion
from scipy.ndimage import distance_transform_edt as np_distance_transform_edt

PM = "\u00B1"
SURFACE_TOL_MM = 3.0
APL_TOL_MM = 0.0

MODE_ORDER = {"inward": 0, "outward": 1, "upshift": 2, "downshift": 3}


# -------------------------
# Fixed evaluation settings
# -------------------------
EVAL_JOBS = [
    {
        "name": "pre_eso",
        "gt_dir": "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_raw/Dataset009_EsoCTV73pAll/labelsTs",
        "pred_root": (
            "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_results/"
            "Dataset008_EsoCTV73p/nnUNetTrainer__nnUNetPlans__3d_fullres/"
            "TestResults_28p_fold1/Pre_crop_error_fullsize"
        ),
        "excel_name": "Eval_pre_error.xlsx",
    },
    {
        "name": "pre_rectal",
        "gt_dir": "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_raw/Dataset011_RectalCTV146p/labelsTs",
        "pred_root": (
            "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_results/"
            "Dataset013_RectalCTV146pCrop/nnUNetTrainer__nnUNetPlans__3d_fullres/"
            "TestResults_fold0/Pre_crop_error_fullsize"
        ),
        "excel_name": "Eval_pre_error.xlsx",
    },
    {
        "name": "post_eso",
        "gt_dir": "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_raw/Dataset009_EsoCTV73pAll/labelsTs",
        "pred_root": (
            "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_results/"
            "Dataset009_EsoCTV73pAll/nnUNetTrainer__nnUNetPlans__3d_fullres/"
            "TestResults_28p_fold4/Post_crop_error_fullsize"
        ),
        "excel_name": "Eval_post_error.xlsx",
    },
    {
        "name": "post_rectal",
        "gt_dir": "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_raw/Dataset011_RectalCTV146p/labelsTs",
        "pred_root": (
            "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_results/"
            "Dataset011_RectalCTV146p/nnUNetTrainer__nnUNetPlans__3d_fullres/"
            "TestResults_fold0/Post_crop_error_fullsize"
        ),
        "excel_name": "Eval_post_error.xlsx",
    },
]


@dataclass
class ArrayBackend:
    name: str
    xp: Any
    binary_erosion: Any
    distance_transform_edt: Any
    use_cuda: bool

    def asarray(self, arr: np.ndarray) -> Any:
        if self.use_cuda:
            return self.xp.asarray(arr)
        return np.asarray(arr)

    def to_numpy(self, arr: Any) -> np.ndarray:
        if self.use_cuda:
            return self.xp.asnumpy(arr)
        return np.asarray(arr)

    def scalar_float(self, value: Any) -> float:
        if self.use_cuda:
            return float(self.xp.asnumpy(value))
        return float(value)

    def scalar_bool(self, value: Any) -> bool:
        if self.use_cuda:
            return bool(self.xp.asnumpy(value))
        return bool(value)


def get_array_backend(device: str) -> ArrayBackend:
    req = device.lower()
    if req not in {"auto", "cpu", "cuda"}:
        raise ValueError("--device must be one of: auto, cpu, cuda")

    if req in {"auto", "cuda"}:
        try:
            import cupy as cp  # type: ignore
            from cupyx.scipy.ndimage import binary_erosion as cp_binary_erosion  # type: ignore
            from cupyx.scipy.ndimage import distance_transform_edt as cp_distance_transform_edt  # type: ignore

            try:
                _ = cp.cuda.runtime.getDeviceCount()
            except Exception as e:
                if req == "cuda":
                    raise RuntimeError(f"CUDA requested but not usable: {e}") from e
            else:
                return ArrayBackend(
                    name="cupy-cuda",
                    xp=cp,
                    binary_erosion=cp_binary_erosion,
                    distance_transform_edt=cp_distance_transform_edt,
                    use_cuda=True,
                )
        except Exception as e:
            if req == "cuda":
                raise RuntimeError(
                    "CUDA requested, but CuPy/cupyx is not available. "
                    "Please install CuPy with CUDA support."
                ) from e

    return ArrayBackend(
        name="numpy-cpu",
        xp=np,
        binary_erosion=np_binary_erosion,
        distance_transform_edt=np_distance_transform_edt,
        use_cuda=False,
    )


def strip_nii_ext(name: str) -> str:
    if name.endswith(".nii.gz"):
        return name[:-7]
    if name.endswith(".nii"):
        return name[:-4]
    return name


def canonical_case_id_from_stem(stem: str) -> str:
    if stem.endswith("_0000"):
        return stem[:-5]
    return stem


def list_nii_files(folder: str) -> List[str]:
    return sorted(
        [f for f in os.listdir(folder) if f.endswith(".nii.gz") or f.endswith(".nii")]
    )


def extract_numeric_key_optional(name: str) -> Optional[str]:
    nums = re.findall(r"\d+", name)
    if not nums:
        return None
    return str(int(nums[-1]))


def build_gt_indices(gt_dir: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    by_case: Dict[str, str] = {}
    by_num: Dict[str, str] = {}

    for fname in list_nii_files(gt_dir):
        path = os.path.join(gt_dir, fname)
        stem = strip_nii_ext(fname)
        cid = canonical_case_id_from_stem(stem)

        if cid in by_case:
            raise ValueError(f"Duplicate GT case id: {cid} in {gt_dir}")
        by_case[cid] = path

        num = extract_numeric_key_optional(cid)
        if num is not None and num not in by_num:
            by_num[num] = path

    return by_case, by_num


def get_error_subdirs(pred_root: str) -> List[str]:
    cands = []
    for name in os.listdir(pred_root):
        p = os.path.join(pred_root, name)
        if os.path.isdir(p):
            cands.append(name)

    def sort_key(n: str) -> Tuple[int, int, str]:
        m = re.match(r"^K(\d+)_(\w+)$", n)
        if not m:
            return (9999, 9999, n)
        k = int(m.group(1))
        mode = m.group(2)
        return (k, MODE_ORDER.get(mode, 9999), n)

    return sorted(cands, key=sort_key)


def spacing_xyz_to_zyx(spacing_xyz: Tuple[float, float, float]) -> Tuple[float, float, float]:
    sx, sy, sz = [float(v) for v in spacing_xyz]
    return sz, sy, sx


def spacing_xyz_to_yx(spacing_xyz: Tuple[float, float, float]) -> Tuple[float, float]:
    sx, sy, _ = [float(v) for v in spacing_xyz]
    return sy, sx


def dice_coefficient(mask_a: Any, mask_b: Any, backend: ArrayBackend) -> float:
    a = mask_a > 0
    b = mask_b > 0
    a_sum = backend.scalar_float(backend.xp.sum(a))
    b_sum = backend.scalar_float(backend.xp.sum(b))

    if a_sum == 0.0 and b_sum == 0.0:
        return float("nan")

    inter = backend.scalar_float(backend.xp.sum(backend.xp.logical_and(a, b)))
    return float((2.0 * inter) / (a_sum + b_sum))


def extract_surface(mask: Any, backend: ArrayBackend) -> Any:
    xp = backend.xp
    mask = mask.astype(bool)
    if not backend.scalar_bool(mask.any()):
        return xp.zeros_like(mask, dtype=bool)

    if mask.ndim == 3:
        structure = xp.zeros((3, 3, 3), dtype=bool)
        structure[1, 1, 1] = True
        structure[0, 1, 1] = True
        structure[2, 1, 1] = True
        structure[1, 0, 1] = True
        structure[1, 2, 1] = True
        structure[1, 1, 0] = True
        structure[1, 1, 2] = True
    elif mask.ndim == 2:
        structure = xp.array(
            [
                [0, 1, 0],
                [1, 1, 1],
                [0, 1, 0],
            ],
            dtype=bool,
        )
    else:
        raise ValueError(f"Unsupported ndim={mask.ndim}")

    eroded = backend.binary_erosion(mask, structure=structure, border_value=0)
    return mask & (~eroded)


def hd95_from_masks(mask_a: Any, mask_b: Any, spacing: Tuple[float, ...], backend: ArrayBackend) -> float:
    a_any = backend.scalar_bool(mask_a.any())
    b_any = backend.scalar_bool(mask_b.any())
    if not a_any or not b_any:
        return float("nan")
    # Keep HD95 fully aligned with evaluate_metrics.py (medpy implementation).
    a_np = backend.to_numpy(mask_a > 0).astype(np.uint8)
    b_np = backend.to_numpy(mask_b > 0).astype(np.uint8)
    return float(metric.binary.hd95(a_np, b_np, voxelspacing=spacing))


def compute_2d_metrics_nonempty_slices(
    pred_zyx: Any,
    gt_zyx: Any,
    spacing_xyz: Tuple[float, float, float],
    backend: ArrayBackend,
) -> Tuple[float, float]:
    xp = backend.xp
    dice_vals: List[float] = []
    hd95_vals: List[float] = []
    spacing_yx = spacing_xyz_to_yx(spacing_xyz)

    z_dim = int(gt_zyx.shape[0])
    for z in range(z_dim):
        gt_slice = gt_zyx[z] > 0
        pred_slice = pred_zyx[z] > 0

        gt_any = backend.scalar_bool(gt_slice.any())
        pred_any = backend.scalar_bool(pred_slice.any())
        if (not gt_any) and (not pred_any):
            continue

        d = dice_coefficient(pred_slice, gt_slice, backend)
        if not math.isnan(d):
            dice_vals.append(d)

        if gt_any and pred_any:
            hd95_vals.append(hd95_from_masks(pred_slice, gt_slice, spacing_yx, backend))

    dice_mean = float(np.mean(dice_vals)) if dice_vals else float("nan")
    hd95_mean = float(np.mean(hd95_vals)) if hd95_vals else float("nan")
    return dice_mean, hd95_mean


def compute_surface_area_weights_3d(mask: Any, spacing_xyz: Tuple[float, float, float], backend: ArrayBackend) -> Any:
    xp = backend.xp
    mask = mask.astype(bool)
    if not backend.scalar_bool(mask.any()):
        return xp.zeros_like(mask, dtype=xp.float64)

    sx, sy, sz = [float(v) for v in spacing_xyz]
    area_x_face = sy * sz
    area_y_face = sx * sz
    area_z_face = sx * sy

    w = xp.zeros(mask.shape, dtype=xp.float64)

    bg = xp.ones_like(mask, dtype=bool)
    bg[1:, :, :] = ~mask[:-1, :, :]
    w += (mask & bg) * area_z_face

    bg = xp.ones_like(mask, dtype=bool)
    bg[:-1, :, :] = ~mask[1:, :, :]
    w += (mask & bg) * area_z_face

    bg = xp.ones_like(mask, dtype=bool)
    bg[:, 1:, :] = ~mask[:, :-1, :]
    w += (mask & bg) * area_y_face

    bg = xp.ones_like(mask, dtype=bool)
    bg[:, :-1, :] = ~mask[:, 1:, :]
    w += (mask & bg) * area_y_face

    bg = xp.ones_like(mask, dtype=bool)
    bg[:, :, 1:] = ~mask[:, :, :-1]
    w += (mask & bg) * area_x_face

    bg = xp.ones_like(mask, dtype=bool)
    bg[:, :, :-1] = ~mask[:, :, 1:]
    w += (mask & bg) * area_x_face

    return w


def compute_surface_dsc_3d(
    pred_zyx: Any,
    gt_zyx: Any,
    spacing_xyz: Tuple[float, float, float],
    backend: ArrayBackend,
    tolerance_mm: float = SURFACE_TOL_MM,
) -> float:
    pred_bin = pred_zyx > 0
    gt_bin = gt_zyx > 0

    pred_any = backend.scalar_bool(pred_bin.any())
    gt_any = backend.scalar_bool(gt_bin.any())

    if (not pred_any) and (not gt_any):
        return 1.0
    if pred_any != gt_any:
        return 0.0

    pred_surface = extract_surface(pred_bin, backend)
    gt_surface = extract_surface(gt_bin, backend)

    pred_weights = compute_surface_area_weights_3d(pred_bin, spacing_xyz, backend) * pred_surface
    gt_weights = compute_surface_area_weights_3d(gt_bin, spacing_xyz, backend) * gt_surface

    pred_surface_any = backend.scalar_bool(pred_surface.any())
    gt_surface_any = backend.scalar_bool(gt_surface.any())
    if (not pred_surface_any) and (not gt_surface_any):
        return 1.0
    if (not pred_surface_any) or (not gt_surface_any):
        return 0.0

    spacing_zyx = spacing_xyz_to_zyx(spacing_xyz)
    dist_to_gt_surface = backend.distance_transform_edt(~gt_surface, sampling=spacing_zyx)
    dist_to_pred_surface = backend.distance_transform_edt(~pred_surface, sampling=spacing_zyx)

    pred_match = pred_surface & (dist_to_gt_surface <= tolerance_mm)
    gt_match = gt_surface & (dist_to_pred_surface <= tolerance_mm)

    matched_area = backend.scalar_float(pred_weights[pred_match].sum() + gt_weights[gt_match].sum())
    total_area = backend.scalar_float(pred_weights[pred_surface].sum() + gt_weights[gt_surface].sum())

    if total_area == 0.0:
        return 1.0
    return float(matched_area / total_area)


def compute_boundary_length_weights_2d(mask_yx: Any, spacing_xyz: Tuple[float, float, float], backend: ArrayBackend) -> Any:
    xp = backend.xp
    mask = mask_yx.astype(bool)
    if not backend.scalar_bool(mask.any()):
        return xp.zeros_like(mask, dtype=xp.float64)

    sx, sy, _ = [float(v) for v in spacing_xyz]
    w = xp.zeros(mask.shape, dtype=xp.float64)

    bg = xp.ones_like(mask, dtype=bool)
    bg[1:, :] = ~mask[:-1, :]
    w += (mask & bg) * sx

    bg = xp.ones_like(mask, dtype=bool)
    bg[:-1, :] = ~mask[1:, :]
    w += (mask & bg) * sx

    bg = xp.ones_like(mask, dtype=bool)
    bg[:, 1:] = ~mask[:, :-1]
    w += (mask & bg) * sy

    bg = xp.ones_like(mask, dtype=bool)
    bg[:, :-1] = ~mask[:, 1:]
    w += (mask & bg) * sy

    return w


def compute_apl_bidirectional_mm(
    pred_zyx: Any,
    gt_zyx: Any,
    spacing_xyz: Tuple[float, float, float],
    backend: ArrayBackend,
    tolerance_mm: float = APL_TOL_MM,
) -> Tuple[float, float, float]:
    apl_add_total = 0.0
    apl_del_total = 0.0
    spacing_yx = spacing_xyz_to_yx(spacing_xyz)

    z_dim = int(gt_zyx.shape[0])
    for z in range(z_dim):
        gt_slice = gt_zyx[z] > 0
        pred_slice = pred_zyx[z] > 0

        gt_any = backend.scalar_bool(gt_slice.any())
        pred_any = backend.scalar_bool(pred_slice.any())
        if (not gt_any) and (not pred_any):
            continue

        gt_boundary = extract_surface(gt_slice, backend)
        pred_boundary = extract_surface(pred_slice, backend)

        gt_weights = compute_boundary_length_weights_2d(gt_slice, spacing_xyz, backend) * gt_boundary
        pred_weights = compute_boundary_length_weights_2d(pred_slice, spacing_xyz, backend) * pred_boundary

        gt_boundary_any = backend.scalar_bool(gt_boundary.any())
        pred_boundary_any = backend.scalar_bool(pred_boundary.any())
        if (not gt_boundary_any) and (not pred_boundary_any):
            continue

        if not pred_boundary_any:
            apl_add_total += backend.scalar_float(gt_weights[gt_boundary].sum())
            continue

        if not gt_boundary_any:
            apl_del_total += backend.scalar_float(pred_weights[pred_boundary].sum())
            continue

        dist_to_pred_boundary = backend.distance_transform_edt(~pred_boundary, sampling=spacing_yx)
        dist_to_gt_boundary = backend.distance_transform_edt(~gt_boundary, sampling=spacing_yx)

        add_mask = gt_boundary & (dist_to_pred_boundary > tolerance_mm)
        del_mask = pred_boundary & (dist_to_gt_boundary > tolerance_mm)

        apl_add_total += backend.scalar_float(gt_weights[add_mask].sum())
        apl_del_total += backend.scalar_float(pred_weights[del_mask].sum())

    apl_total = apl_add_total + apl_del_total
    return apl_add_total, apl_del_total, apl_total


def tuple_close(a: Tuple[float, ...], b: Tuple[float, ...], tol: float = 1e-6) -> bool:
    if len(a) != len(b):
        return False
    return all(abs(float(x) - float(y)) <= tol for x, y in zip(a, b))


def check_image_consistency(gt_img: sitk.Image, pred_img: sitk.Image) -> Tuple[bool, str]:
    if gt_img.GetSize() != pred_img.GetSize():
        return False, f"size mismatch: gt={gt_img.GetSize()}, pred={pred_img.GetSize()}"
    if not tuple_close(gt_img.GetSpacing(), pred_img.GetSpacing()):
        return False, f"spacing mismatch: gt={gt_img.GetSpacing()}, pred={pred_img.GetSpacing()}"
    if not tuple_close(gt_img.GetOrigin(), pred_img.GetOrigin()):
        return False, f"origin mismatch: gt={gt_img.GetOrigin()}, pred={pred_img.GetOrigin()}"
    if not tuple_close(gt_img.GetDirection(), pred_img.GetDirection()):
        return False, f"direction mismatch: gt={gt_img.GetDirection()}, pred={pred_img.GetDirection()}"
    return True, ""


def auto_fit_and_center(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def write_model_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    columns: List[str],
    rows: List[List[object]],
    mean_row: List[object],
    std_row: List[object],
) -> None:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append(row)
    ws.append(mean_row)
    ws.append(std_row)

    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float, np.floating)):
                cell.number_format = "0.00"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    auto_fit_and_center(ws)


def generate_summary_mean_sheet(wb: openpyxl.Workbook) -> None:
    metric_cols = {
        "2D Dice",
        "2D HD95 (mm)",
        "3D Dice",
        "3D HD95 (mm)",
        f"Surface DSC ({int(SURFACE_TOL_MM)}mm)",
        "APL_add (mm)",
        "APL_del (mm)",
        "APL_total (mm)",
    }

    if "Summary_Mean" in wb.sheetnames:
        wb.remove(wb["Summary_Mean"])
    ws_sum = wb.create_sheet("Summary_Mean", 0)

    ref_columns: List[str] = []
    summary_rows: List[List[object]] = []

    for sheet in wb.sheetnames:
        if sheet == "Summary_Mean":
            continue

        ws = wb[sheet]
        if ws.max_row < 3:
            continue

        if not ref_columns:
            ref_columns = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

        mean_vals = [ws.cell(row=ws.max_row - 1, column=i).value for i in range(1, ws.max_column + 1)]
        std_vals = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]

        combined = [sheet]
        for i, col in enumerate(ref_columns[1:], start=1):
            mv, sv = mean_vals[i], std_vals[i]
            if col in metric_cols:
                try:
                    combined.append(f"{float(mv):.2f} {PM} {float(sv):.2f}")
                except Exception:
                    combined.append("")
            else:
                combined.append(mv if mv is not None else "")
        summary_rows.append(combined)

    if not ref_columns:
        ws_sum.append(["Error"])
        auto_fit_and_center(ws_sum)
        return

    ws_sum.append(["Error"] + ref_columns[1:])
    for row in summary_rows:
        ws_sum.append(row)

    for cell in ws_sum[1]:
        cell.font = Font(bold=True)

    auto_fit_and_center(ws_sum)


def find_gt_path_for_pred(
    pred_name: str,
    gt_by_case: Dict[str, str],
    gt_by_num: Dict[str, str],
) -> Tuple[str, str]:
    stem = strip_nii_ext(pred_name)
    cid = canonical_case_id_from_stem(stem)

    p = gt_by_case.get(cid, "")
    if p:
        return cid, p

    num = extract_numeric_key_optional(cid)
    if num is not None and num in gt_by_num:
        return cid, gt_by_num[num]

    return cid, ""


def evaluate_one_error_sheet(
    wb: openpyxl.Workbook,
    error_name: str,
    pred_dir: str,
    gt_by_case: Dict[str, str],
    gt_by_num: Dict[str, str],
    backend: ArrayBackend,
) -> None:
    print(f"\nEvaluating error: {error_name}")

    columns = [
        "ID",
        "2D Dice",
        "2D HD95 (mm)",
        "3D Dice",
        "3D HD95 (mm)",
        f"Surface DSC ({int(SURFACE_TOL_MM)}mm)",
        "APL_add (mm)",
        "APL_del (mm)",
        "APL_total (mm)",
    ]

    pred_files = list_nii_files(pred_dir)
    rows: List[List[object]] = []

    for pred_name in pred_files:
        pred_path = os.path.join(pred_dir, pred_name)
        case_id, gt_path = find_gt_path_for_pred(pred_name, gt_by_case, gt_by_num)

        if not gt_path:
            print(f"[Skip] GT not found for pred={pred_name}")
            rows.append([case_id, "", "", "", "", "", "", "", ""])
            continue

        gt_img = sitk.ReadImage(gt_path)
        pred_img = sitk.ReadImage(pred_path)

        ok, msg = check_image_consistency(gt_img, pred_img)
        if not ok:
            print(f"[Skip] {case_id}: {msg}")
            rows.append([case_id, "", "", "", "", "", "", "", ""])
            continue

        gt_np = sitk.GetArrayFromImage(gt_img).astype(np.uint8)  # [Z, Y, X]
        pred_np = sitk.GetArrayFromImage(pred_img).astype(np.uint8)
        spacing_xyz = gt_img.GetSpacing()

        gt = backend.asarray(gt_np)
        pred = backend.asarray(pred_np)

        dice2d, hd95_2d = compute_2d_metrics_nonempty_slices(pred, gt, spacing_xyz, backend)
        dice3d = dice_coefficient(pred, gt, backend)
        hd95_3d = hd95_from_masks(pred > 0, gt > 0, spacing_xyz_to_zyx(spacing_xyz), backend)
        surface_dsc = compute_surface_dsc_3d(pred, gt, spacing_xyz, backend, tolerance_mm=SURFACE_TOL_MM)
        apl_add_mm, apl_del_mm, apl_total_mm = compute_apl_bidirectional_mm(
            pred, gt, spacing_xyz, backend, tolerance_mm=APL_TOL_MM
        )

        rows.append(
            [
                case_id,
                round(dice2d, 2) if not np.isnan(dice2d) else "",
                round(hd95_2d, 2) if not np.isnan(hd95_2d) else "",
                round(dice3d, 2) if not np.isnan(dice3d) else "",
                round(hd95_3d, 2) if not np.isnan(hd95_3d) else "",
                round(surface_dsc, 2) if not np.isnan(surface_dsc) else "",
                round(apl_add_mm, 2),
                round(apl_del_mm, 2),
                round(apl_total_mm, 2),
            ]
        )

    numeric_matrix = []
    for r in rows:
        vals = []
        for v in r[1:]:
            if isinstance(v, (int, float, np.floating)):
                vals.append(float(v))
            else:
                vals.append(np.nan)
        numeric_matrix.append(vals)

    arr = np.asarray(numeric_matrix, dtype=float) if numeric_matrix else np.empty((0, 8))
    mean_row: List[object] = ["Mean"]
    std_row: List[object] = ["STD"]

    if arr.size == 0:
        mean_row.extend([""] * 8)
        std_row.extend([""] * 8)
    else:
        means = np.nanmean(arr, axis=0)
        stds = np.nanstd(arr, axis=0, ddof=1)
        for i in range(8):
            mean_row.append(round(float(means[i]), 2) if not np.isnan(means[i]) else "")
            std_row.append(round(float(stds[i]), 2) if not np.isnan(stds[i]) else "")

    write_model_sheet(wb, error_name, columns, rows, mean_row, std_row)
    print(
        f"[{error_name}] "
        f"2D Dice={mean_row[1]} {PM} {std_row[1]} | "
        f"2D HD95={mean_row[2]} {PM} {std_row[2]} | "
        f"3D Dice={mean_row[3]} {PM} {std_row[3]} | "
        f"3D HD95={mean_row[4]} {PM} {std_row[4]} | "
        f"Surface DSC={mean_row[5]} {PM} {std_row[5]} | "
        f"APL_add={mean_row[6]} {PM} {std_row[6]} | "
        f"APL_del={mean_row[7]} {PM} {std_row[7]} | "
        f"APL_total={mean_row[8]} {PM} {std_row[8]}"
    )


def evaluate_one_root(job: Dict[str, str], backend: ArrayBackend) -> None:
    gt_dir = job["gt_dir"]
    pred_root = job["pred_root"]
    output_excel = os.path.join(pred_root, job["excel_name"])

    if not os.path.isdir(gt_dir):
        print(f"[Skip Job] GT folder not found: {gt_dir}")
        return
    if not os.path.isdir(pred_root):
        print(f"[Skip Job] Prediction root not found: {pred_root}")
        return

    error_dirs = get_error_subdirs(pred_root)
    if not error_dirs:
        print(f"[Skip Job] No error subfolders in: {pred_root}")
        return

    print(f"\n=== Evaluating {job['name']} ===")
    print(f"GT: {gt_dir}")
    print(f"Pred root: {pred_root}")
    print(f"Errors found: {len(error_dirs)}")

    gt_by_case, gt_by_num = build_gt_indices(gt_dir)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    evaluated = 0
    for err in error_dirs:
        pred_dir = os.path.join(pred_root, err)
        if not os.path.isdir(pred_dir):
            continue
        if len(list_nii_files(pred_dir)) == 0:
            print(f"[Skip Error] Empty: {pred_dir}")
            continue
        evaluate_one_error_sheet(wb, err, pred_dir, gt_by_case, gt_by_num, backend)
        evaluated += 1

    if evaluated == 0:
        print(f"[Skip Job] No valid error folders evaluated: {pred_root}")
        return

    generate_summary_mean_sheet(wb)
    wb.save(output_excel)
    print(f"Saved: {output_excel}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate crop-error fullsize predictions.")
    parser.add_argument(
        "--device",
        type=str,
        default="auto",
        choices=["auto", "cpu", "cuda"],
        help="Array backend device for metric computation.",
    )
    parser.add_argument(
        "--which",
        type=str,
        default="all",
        choices=["all", "pre", "post"],
        help="Evaluate all jobs, only pre jobs, or only post jobs.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    backend = get_array_backend(args.device)
    print(f"Backend: {backend.name}")

    if args.which == "pre":
        jobs = [j for j in EVAL_JOBS if j["excel_name"] == "Eval_pre_error.xlsx"]
    elif args.which == "post":
        jobs = [j for j in EVAL_JOBS if j["excel_name"] == "Eval_post_error.xlsx"]
    else:
        jobs = EVAL_JOBS

    for job in jobs:
        evaluate_one_root(job, backend)

    print("\nAll evaluation jobs finished.")


if __name__ == "__main__":
    main()
