"""
Evaluate three nnUNet result folders and save metrics to Excel.

Metrics:
- 2D Dice / 2D HD95
  * Computed on slices where GT > 0 OR Pred > 0
  * 2D HD95 uses correct in-plane spacing (y, x)
  * 2D HD95 is computed only when both GT and Pred are non-empty on the slice

- 3D Dice / 3D HD95
  * 3D HD95 uses correct voxel spacing (z, y, x)
  * 3D HD95 is computed on whole volume and only when both GT and Pred are non-empty

- Surface DSC (3 mm tolerance)
  * 3D surface Dice using physical spacing
  * Surface contribution is weighted by exposed physical surface area

- APL_total (mm)
  * Slice-wise bidirectional contour correction length
  * APL_add: GT contour length outside Pred contour tolerance band
  * APL_del: Pred contour length outside GT contour tolerance band
  * APL_total = sum(APL_add + APL_del) over axial slices
  * Default APL tolerance = 0 mm (classic correction-length style)

Excel format:
- one sheet per model
- case rows + Mean + STD
- Summary_Mean sheet
"""

import math
import os
import re
from typing import Dict, List, Tuple

import numpy as np
import openpyxl
import SimpleITK as sitk
from medpy import metric
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from scipy.ndimage import binary_erosion, distance_transform_edt

PM = "\u00B1"
SURFACE_TOL_MM = 3.0
APL_TOL_MM = 0.0  # set >0 if you want tolerance-based APL


# =========================
# Basic file / ID utilities
# =========================
def strip_nii_ext(name: str) -> str:
    if name.endswith(".nii.gz"):
        return name[:-7]
    if name.endswith(".nii"):
        return name[:-4]
    return name


def extract_numeric_key(name: str) -> str:
    stem = strip_nii_ext(os.path.basename(name))
    nums = re.findall(r"\d+", stem)
    if not nums:
        raise ValueError(f"No numeric id found in filename: {name}")
    return str(int(nums[-1]))


def list_nii_files(folder: str) -> List[str]:
    return sorted([f for f in os.listdir(folder) if f.endswith(".nii.gz") or f.endswith(".nii")])


def build_numeric_index(folder: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for name in list_nii_files(folder):
        key = extract_numeric_key(name)
        path = os.path.join(folder, name)
        if key in out:
            raise ValueError(
                f"Duplicate numeric id in folder {folder}: id={key}, "
                f"files={os.path.basename(out[key])} and {name}"
            )
        out[key] = path
    return out


# =========================
# Basic metric utilities
# =========================
def dice_coefficient(a: np.ndarray, b: np.ndarray) -> float:
    a = a > 0
    b = b > 0
    a_sum = int(a.sum())
    b_sum = int(b.sum())

    if a_sum == 0 and b_sum == 0:
        return float("nan")

    inter = np.logical_and(a, b).sum()
    return float((2.0 * inter) / (a_sum + b_sum))


def spacing_xyz_to_zyx(spacing_xyz: Tuple[float, float, float]) -> Tuple[float, float, float]:
    sx, sy, sz = [float(v) for v in spacing_xyz]
    return sz, sy, sx


def spacing_xyz_to_yx(spacing_xyz: Tuple[float, float, float]) -> Tuple[float, float]:
    sx, sy, _ = [float(v) for v in spacing_xyz]
    return sy, sx


# =========================
# HD95
# =========================
def safe_hd95_3d(pred: np.ndarray, gt: np.ndarray, spacing_xyz: Tuple[float, float, float]) -> float:
    pred_bin = (pred > 0).astype(np.uint8)
    gt_bin = (gt > 0).astype(np.uint8)

    pred_any = pred_bin.max() > 0
    gt_any = gt_bin.max() > 0

    if pred_any and gt_any:
        return float(metric.binary.hd95(pred_bin, gt_bin, voxelspacing=spacing_xyz_to_zyx(spacing_xyz)))

    return float("nan")


def compute_2d_metrics_nonempty_slices(
    pred_zyx: np.ndarray,
    gt_zyx: np.ndarray,
    spacing_xyz: Tuple[float, float, float],
) -> Tuple[float, float]:
    """
    2D Dice:
        computed on slices where GT > 0 OR Pred > 0
    2D HD95:
        computed only on slices where both GT and Pred are non-empty
    """
    dice_vals: List[float] = []
    hd95_vals: List[float] = []

    spacing_yx = spacing_xyz_to_yx(spacing_xyz)

    for z in range(gt_zyx.shape[0]):
        gt_slice = (gt_zyx[z] > 0).astype(np.uint8)
        pred_slice = (pred_zyx[z] > 0).astype(np.uint8)

        # Skip double-empty slices
        if gt_slice.max() == 0 and pred_slice.max() == 0:
            continue

        d = dice_coefficient(pred_slice, gt_slice)
        if not np.isnan(d):
            dice_vals.append(d)

        if gt_slice.max() > 0 and pred_slice.max() > 0:
            hd95_vals.append(float(metric.binary.hd95(pred_slice, gt_slice, voxelspacing=spacing_yx)))

    dice_mean = float(np.mean(dice_vals)) if dice_vals else float("nan")
    hd95_mean = float(np.mean(hd95_vals)) if hd95_vals else float("nan")
    return dice_mean, hd95_mean


# =========================
# Surface DSC (area-weighted)
# =========================
def extract_surface_3d(mask: np.ndarray) -> np.ndarray:
    """
    6-neighborhood surface voxels.
    """
    mask = mask.astype(bool)
    if not mask.any():
        return np.zeros_like(mask, dtype=bool)

    structure = np.zeros((3, 3, 3), dtype=bool)
    structure[1, 1, 1] = True
    structure[0, 1, 1] = True
    structure[2, 1, 1] = True
    structure[1, 0, 1] = True
    structure[1, 2, 1] = True
    structure[1, 1, 0] = True
    structure[1, 1, 2] = True

    eroded = binary_erosion(mask, structure=structure, border_value=0)
    return mask & (~eroded)


def compute_surface_area_weights_3d(mask: np.ndarray, spacing_xyz: Tuple[float, float, float]) -> np.ndarray:
    """
    For each foreground voxel, compute the physical area of its exposed faces.
    This gives a surface-area weight map for surface voxels.

    Axes in array are [Z, Y, X]
    Spacing from SimpleITK is (X, Y, Z)
    """
    mask = mask.astype(bool)
    if not mask.any():
        return np.zeros_like(mask, dtype=np.float64)

    sx, sy, sz = [float(v) for v in spacing_xyz]

    # Face areas
    area_x_face = sy * sz  # face normal along x
    area_y_face = sx * sz  # face normal along y
    area_z_face = sx * sy  # face normal along z

    w = np.zeros(mask.shape, dtype=np.float64)

    # negative z exposed face
    bg = np.ones_like(mask, dtype=bool)
    bg[1:, :, :] = ~mask[:-1, :, :]
    w += (mask & bg) * area_z_face

    # positive z exposed face
    bg = np.ones_like(mask, dtype=bool)
    bg[:-1, :, :] = ~mask[1:, :, :]
    w += (mask & bg) * area_z_face

    # negative y exposed face
    bg = np.ones_like(mask, dtype=bool)
    bg[:, 1:, :] = ~mask[:, :-1, :]
    w += (mask & bg) * area_y_face

    # positive y exposed face
    bg = np.ones_like(mask, dtype=bool)
    bg[:, :-1, :] = ~mask[:, 1:, :]
    w += (mask & bg) * area_y_face

    # negative x exposed face
    bg = np.ones_like(mask, dtype=bool)
    bg[:, :, 1:] = ~mask[:, :, :-1]
    w += (mask & bg) * area_x_face

    # positive x exposed face
    bg = np.ones_like(mask, dtype=bool)
    bg[:, :, :-1] = ~mask[:, :, 1:]
    w += (mask & bg) * area_x_face

    return w


def compute_surface_dsc_3d(
    pred_zyx: np.ndarray,
    gt_zyx: np.ndarray,
    spacing_xyz: Tuple[float, float, float],
    tolerance_mm: float = SURFACE_TOL_MM,
) -> float:
    """
    Area-weighted surface DSC.

    A surface voxel contributes its exposed physical surface area.
    A surface element is counted as matched if its nearest distance
    to the opposite surface is <= tolerance_mm.
    """
    pred_bin = pred_zyx > 0
    gt_bin = gt_zyx > 0

    if not pred_bin.any() and not gt_bin.any():
        return 1.0
    if pred_bin.any() != gt_bin.any():
        return 0.0

    pred_surface = extract_surface_3d(pred_bin)
    gt_surface = extract_surface_3d(gt_bin)

    pred_weights = compute_surface_area_weights_3d(pred_bin, spacing_xyz) * pred_surface
    gt_weights = compute_surface_area_weights_3d(gt_bin, spacing_xyz) * gt_surface

    if pred_surface.sum() == 0 and gt_surface.sum() == 0:
        return 1.0
    if pred_surface.sum() == 0 or gt_surface.sum() == 0:
        return 0.0

    spacing_zyx = spacing_xyz_to_zyx(spacing_xyz)

    # Distance from each voxel center to nearest surface voxel center
    dist_to_gt_surface = distance_transform_edt(~gt_surface, sampling=spacing_zyx)
    dist_to_pred_surface = distance_transform_edt(~pred_surface, sampling=spacing_zyx)

    pred_match = pred_surface & (dist_to_gt_surface <= tolerance_mm)
    gt_match = gt_surface & (dist_to_pred_surface <= tolerance_mm)

    matched_area = float(pred_weights[pred_match].sum() + gt_weights[gt_match].sum())
    total_area = float(pred_weights[pred_surface].sum() + gt_weights[gt_surface].sum())

    if total_area == 0:
        return 1.0
    return matched_area / total_area


# =========================
# APL (standardized contour-correction style)
# =========================
def extract_boundary_2d(mask_yx: np.ndarray) -> np.ndarray:
    """
    4-neighborhood inner boundary pixels.
    """
    mask = mask_yx.astype(bool)
    if not mask.any():
        return np.zeros_like(mask, dtype=bool)

    structure = np.array(
        [
            [0, 1, 0],
            [1, 1, 1],
            [0, 1, 0],
        ],
        dtype=bool,
    )
    eroded = binary_erosion(mask, structure=structure, border_value=0)
    return mask & (~eroded)


def compute_boundary_length_weights_2d(mask_yx: np.ndarray, spacing_xyz: Tuple[float, float, float]) -> np.ndarray:
    """
    For each foreground pixel, compute the physical length of its exposed contour edges.

    Array axes: [Y, X]
    Spacing from SimpleITK: (X, Y, Z)
    """
    mask = mask_yx.astype(bool)
    if not mask.any():
        return np.zeros_like(mask, dtype=np.float64)

    sx, sy, _ = [float(v) for v in spacing_xyz]

    w = np.zeros(mask.shape, dtype=np.float64)

    # negative y exposed edge
    bg = np.ones_like(mask, dtype=bool)
    bg[1:, :] = ~mask[:-1, :]
    w += (mask & bg) * sx

    # positive y exposed edge
    bg = np.ones_like(mask, dtype=bool)
    bg[:-1, :] = ~mask[1:, :]
    w += (mask & bg) * sx

    # negative x exposed edge
    bg = np.ones_like(mask, dtype=bool)
    bg[:, 1:] = ~mask[:, :-1]
    w += (mask & bg) * sy

    # positive x exposed edge
    bg = np.ones_like(mask, dtype=bool)
    bg[:, :-1] = ~mask[:, 1:]
    w += (mask & bg) * sy

    return w


def compute_apl_bidirectional_mm(
    pred_zyx: np.ndarray,
    gt_zyx: np.ndarray,
    spacing_xyz: Tuple[float, float, float],
    tolerance_mm: float = APL_TOL_MM,
) -> Tuple[float, float, float]:
    """
    Bidirectional contour correction length, slice-wise.

    APL_add:
        GT contour length farther than tolerance_mm from Pred contour
        -> contour that must be added

    APL_del:
        Pred contour length farther than tolerance_mm from GT contour
        -> contour that must be deleted

    APL_total = APL_add + APL_del
    """
    apl_add_total = 0.0
    apl_del_total = 0.0
    spacing_yx = spacing_xyz_to_yx(spacing_xyz)

    for z in range(gt_zyx.shape[0]):
        gt_slice = gt_zyx[z] > 0
        pred_slice = pred_zyx[z] > 0

        if not gt_slice.any() and not pred_slice.any():
            continue

        gt_boundary = extract_boundary_2d(gt_slice)
        pred_boundary = extract_boundary_2d(pred_slice)

        gt_weights = compute_boundary_length_weights_2d(gt_slice, spacing_xyz) * gt_boundary
        pred_weights = compute_boundary_length_weights_2d(pred_slice, spacing_xyz) * pred_boundary

        # Handle one-sided empty contour cases
        if gt_boundary.sum() == 0 and pred_boundary.sum() == 0:
            continue

        if pred_boundary.sum() == 0:
            apl_add_total += float(gt_weights[gt_boundary].sum())
            continue

        if gt_boundary.sum() == 0:
            apl_del_total += float(pred_weights[pred_boundary].sum())
            continue

        dist_to_pred_boundary = distance_transform_edt(~pred_boundary, sampling=spacing_yx)
        dist_to_gt_boundary = distance_transform_edt(~gt_boundary, sampling=spacing_yx)

        add_mask = gt_boundary & (dist_to_pred_boundary > tolerance_mm)
        del_mask = pred_boundary & (dist_to_gt_boundary > tolerance_mm)

        apl_add_total += float(gt_weights[add_mask].sum())
        apl_del_total += float(pred_weights[del_mask].sum())

    apl_total = apl_add_total + apl_del_total
    return apl_add_total, apl_del_total, apl_total


# =========================
# Image consistency checks
# =========================
def tuple_close(a: Tuple[float, ...], b: Tuple[float, ...], tol: float = 1e-6) -> bool:
    if len(a) != len(b):
        return False
    return all(abs(float(x) - float(y)) <= tol for x, y in zip(a, b))


def check_image_consistency(gt_img: sitk.Image, pred_img: sitk.Image) -> Tuple[bool, str]:
    if gt_img.GetSize() != pred_img.GetSize():
        return False, f"size mismatch: gt={gt_img.GetSize()}, pred={pred_img.GetSize()}"
    if not tuple_close(gt_img.GetSpacing(), pred_img.GetSpacing()):
        return False, f"spacing mismatch: gt={gt_img.GetSpacing()}, pred={pred_img.GetSpacing()}"
    if not tuple_close(gt_img.GetOrigin(), pred_img.GetOrigin()):
        return False, f"origin mismatch: gt={gt_img.GetOrigin()}, pred={pred_img.GetOrigin()}"
    if not tuple_close(gt_img.GetDirection(), pred_img.GetDirection()):
        return False, f"direction mismatch: gt={gt_img.GetDirection()}, pred={pred_img.GetDirection()}"
    return True, ""


# =========================
# Excel helpers
# =========================
def auto_fit_and_center(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def write_model_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    columns: List[str],
    rows: List[List[object]],
    mean_row: List[object],
    std_row: List[object],
) -> None:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append(row)
    ws.append(mean_row)
    ws.append(std_row)

    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float, np.floating)):
                cell.number_format = "0.00"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    auto_fit_and_center(ws)


# =========================
# Main evaluation
# =========================
def evaluate_one_model(
    wb: openpyxl.Workbook,
    model_name: str,
    pred_dir: str,
    gt_index: Dict[str, str],
) -> None:
    print(f"\nEvaluating model: {model_name}")
    pred_index = build_numeric_index(pred_dir)

    columns = [
        "ID",
        "2D Dice",
        "2D HD95 (mm)",
        "3D Dice",
        "3D HD95 (mm)",
        f"Surface DSC ({int(SURFACE_TOL_MM)}mm)",
        "APL_add (mm)",
        "APL_del (mm)",
        "APL_total (mm)",
    ]

    rows: List[List[object]] = []
    keys = sorted(pred_index.keys(), key=lambda x: int(x))

    for key in keys:
        pred_path = pred_index[key]
        case_id = f"p_{int(key)}"
        gt_path = gt_index.get(key, "")

        if not gt_path:
            print(f"[Skip] GT not found for id={key}, pred={os.path.basename(pred_path)}")
            rows.append([case_id, "", "", "", "", "", "", "", ""])
            continue

        gt_img = sitk.ReadImage(gt_path)
        pred_img = sitk.ReadImage(pred_path)

        ok, msg = check_image_consistency(gt_img, pred_img)
        if not ok:
            print(f"[Skip] {case_id}: {msg}")
            rows.append([case_id, "", "", "", "", "", "", "", ""])
            continue

        gt = sitk.GetArrayFromImage(gt_img)   # [Z, Y, X]
        pred = sitk.GetArrayFromImage(pred_img)
        spacing_xyz = gt_img.GetSpacing()

        dice2d, hd95_2d = compute_2d_metrics_nonempty_slices(pred, gt, spacing_xyz)
        dice3d = dice_coefficient(pred, gt)
        hd95_3d = safe_hd95_3d(pred, gt, spacing_xyz)
        surface_dsc = compute_surface_dsc_3d(pred, gt, spacing_xyz, tolerance_mm=SURFACE_TOL_MM)
        apl_add_mm, apl_del_mm, apl_total_mm = compute_apl_bidirectional_mm(
            pred, gt, spacing_xyz, tolerance_mm=APL_TOL_MM
        )

        rows.append(
            [
                case_id,
                round(dice2d, 2) if not np.isnan(dice2d) else "",
                round(hd95_2d, 2) if not np.isnan(hd95_2d) else "",
                round(dice3d, 2) if not np.isnan(dice3d) else "",
                round(hd95_3d, 2) if not np.isnan(hd95_3d) else "",
                round(surface_dsc, 2) if not np.isnan(surface_dsc) else "",
                round(apl_add_mm, 2),
                round(apl_del_mm, 2),
                round(apl_total_mm, 2),
            ]
        )

    numeric_matrix = []
    for r in rows:
        metric_vals = []
        for v in r[1:]:
            if isinstance(v, (int, float, np.floating)):
                metric_vals.append(float(v))
            else:
                metric_vals.append(np.nan)
        numeric_matrix.append(metric_vals)

    arr = np.asarray(numeric_matrix, dtype=float) if numeric_matrix else np.empty((0, 8))

    mean_row: List[object] = ["Mean"]
    std_row: List[object] = ["STD"]

    if arr.size == 0:
        mean_row.extend([""] * 8)
        std_row.extend([""] * 8)
    else:
        means = np.nanmean(arr, axis=0)
        stds = np.nanstd(arr, axis=0, ddof=1)
        for i in range(8):
            mean_row.append(round(float(means[i]), 2) if not np.isnan(means[i]) else "")
            std_row.append(round(float(stds[i]), 2) if not np.isnan(stds[i]) else "")

    write_model_sheet(wb, model_name, columns, rows, mean_row, std_row)

    print(
        f"[{model_name}] "
        f"2D Dice={mean_row[1]} {PM} {std_row[1]} | "
        f"2D HD95={mean_row[2]} {PM} {std_row[2]} | "
        f"3D Dice={mean_row[3]} {PM} {std_row[3]} | "
        f"3D HD95={mean_row[4]} {PM} {std_row[4]} | "
        f"Surface DSC={mean_row[5]} {PM} {std_row[5]} | "
        f"APL_add={mean_row[6]} {PM} {std_row[6]} | "
        f"APL_del={mean_row[7]} {PM} {std_row[7]} | "
        f"APL_total={mean_row[8]} {PM} {std_row[8]}"
    )


def generate_summary_mean_sheet(wb: openpyxl.Workbook) -> None:
    metric_cols = {
        "2D Dice",
        "2D HD95 (mm)",
        "3D Dice",
        "3D HD95 (mm)",
        f"Surface DSC ({int(SURFACE_TOL_MM)}mm)",
        "APL_add (mm)",
        "APL_del (mm)",
        "APL_total (mm)",
    }

    if "Summary_Mean" in wb.sheetnames:
        wb.remove(wb["Summary_Mean"])
    ws_sum = wb.create_sheet("Summary_Mean", 0)

    ref_columns: List[str] = []
    summary_rows: List[List[object]] = []

    for sheet in wb.sheetnames:
        if sheet == "Summary_Mean":
            continue

        ws = wb[sheet]
        if ws.max_row < 3:
            continue

        if not ref_columns:
            ref_columns = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

        mean_vals = [ws.cell(row=ws.max_row - 1, column=i).value for i in range(1, ws.max_column + 1)]
        std_vals = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]

        combined = [sheet]
        for i, col in enumerate(ref_columns[1:], start=1):
            mv, sv = mean_vals[i], std_vals[i]
            if col in metric_cols:
                try:
                    combined.append(f"{float(mv):.2f} {PM} {float(sv):.2f}")
                except Exception:
                    combined.append("")
            else:
                combined.append(mv if mv is not None else "")
        summary_rows.append(combined)

    if not ref_columns:
        ws_sum.append(["Model"])
        auto_fit_and_center(ws_sum)
        return

    ws_sum.append(["Model"] + ref_columns[1:])
    for row in summary_rows:
        ws_sum.append(row)

    for cell in ws_sum[1]:
        cell.font = Font(bold=True)

    auto_fit_and_center(ws_sum)


# =========================
# Run
# =========================
if __name__ == "__main__":
    base_dir = r"C:\Users\dell\Desktop\Eso_83"
    gt_dir = os.path.join(base_dir, "labelsTs")
    output_excel = os.path.join(base_dir, "Eval_metrics.xlsx")

    model_paths = {
        "nnUNet_baseline": os.path.join(base_dir, "nnUNet_all_volume"),
        "nnUNet_post": os.path.join(base_dir, "nnUNet_all_postprocess"),
        "nnUNet_pre": os.path.join(base_dir, "nnUNet_crop_restore"),
    }

    gt_index = build_numeric_index(gt_dir)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    evaluated = 0
    for model_name, pred_dir in model_paths.items():
        if not os.path.isdir(pred_dir):
            print(f"[Skip Model] Folder not found: {pred_dir}")
            continue
        evaluate_one_model(wb, model_name, pred_dir, gt_index)
        evaluated += 1

    if evaluated == 0:
        print("No model evaluated. Excel not generated.")
    else:
        generate_summary_mean_sheet(wb)
        wb.save(output_excel)
        print(f"\nAll done. Saved to: {output_excel}")