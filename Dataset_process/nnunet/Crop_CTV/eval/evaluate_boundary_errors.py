"""
Evaluate upper/lower boundary layer errors for multiple nnUNet prediction folders.

Per case columns:
- ID
- GT Upper Z
- GT Lower Z (smaller)
- Pred Upper Z
- Pred Lower Z
- Upper Extra Layers
- Upper Missed Layers
- Lower Extra Layers
- Lower Missed Layers

Excel output:
- one sheet per model
"""

import os
import re
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import SimpleITK as sitk
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def strip_nii_ext(name: str) -> str:
    if name.endswith(".nii.gz"):
        return name[:-7]
    if name.endswith(".nii"):
        return name[:-4]
    return name


def extract_numeric_key(name: str) -> str:
    stem = strip_nii_ext(os.path.basename(name))
    nums = re.findall(r"\d+", stem)
    if not nums:
        raise ValueError(f"No numeric id found in filename: {name}")
    return str(int(nums[-1]))


def list_nii_files(folder: str) -> List[str]:
    return sorted([f for f in os.listdir(folder) if f.endswith(".nii.gz") or f.endswith(".nii")])


def build_numeric_index(folder: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for name in list_nii_files(folder):
        key = extract_numeric_key(name)
        path = os.path.join(folder, name)
        if key in out:
            raise ValueError(
                f"Duplicate numeric id in folder {folder}: id={key}, "
                f"files={os.path.basename(out[key])} and {name}"
            )
        out[key] = path
    return out


def auto_fit_and_center(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def get_bounds_z(mask_zyx: np.ndarray) -> Tuple[Optional[int], Optional[int]]:
    z_nonempty = [z for z in range(mask_zyx.shape[0]) if np.any(mask_zyx[z] > 0)]
    if not z_nonempty:
        return None, None
    lower = int(min(z_nonempty))
    upper = int(max(z_nonempty))
    return upper, lower


def compute_boundary_errors(
    gt_upper: Optional[int],
    gt_lower: Optional[int],
    pred_upper: Optional[int],
    pred_lower: Optional[int],
) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    if gt_upper is None or gt_lower is None:
        return None, None, None, None
    if pred_upper is None or pred_lower is None:
        # Prediction empty: boundary errors are undefined in this definition.
        return None, None, None, None

    upper_extra = max(0, pred_upper - gt_upper)
    upper_missed = max(0, gt_upper - pred_upper)

    # Lower index is smaller. Extra means pred extends to smaller z than GT.
    lower_extra = max(0, gt_lower - pred_lower)
    lower_missed = max(0, pred_lower - gt_lower)

    return upper_extra, upper_missed, lower_extra, lower_missed


def evaluate_one_model(
    wb: openpyxl.Workbook,
    model_name: str,
    pred_dir: str,
    gt_index: Dict[str, str],
) -> Dict[str, int]:
    print(f"\nEvaluating model: {model_name}")
    pred_index = build_numeric_index(pred_dir)

    columns = [
        "ID",
        "GT Upper Z",
        "GT Lower Z",
        "Pred Upper Z",
        "Pred Lower Z",
        "Upper Extra Layers",
        "Upper Missed Layers",
        "Lower Extra Layers",
        "Lower Missed Layers",
    ]
    rows: List[List[object]] = []
    valid_rows = 0

    keys = sorted(pred_index.keys(), key=lambda x: int(x))
    for key in keys:
        pred_path = pred_index[key]
        case_id = f"p_{int(key)}"
        gt_path = gt_index.get(key, "")

        if not gt_path:
            print(f"[Skip] GT not found for id={key}, pred={os.path.basename(pred_path)}")
            rows.append([case_id, "", "", "", "", "", "", "", ""])
            continue

        gt_img = sitk.ReadImage(gt_path)
        pred_img = sitk.ReadImage(pred_path)
        gt = sitk.GetArrayFromImage(gt_img)
        pred = sitk.GetArrayFromImage(pred_img)

        if gt.shape != pred.shape:
            print(
                f"[Skip] Shape mismatch id={key}: gt={gt.shape}, pred={pred.shape}, file={os.path.basename(pred_path)}"
            )
            rows.append([case_id, "", "", "", "", "", "", "", ""])
            continue

        gt_upper, gt_lower = get_bounds_z(gt)
        pred_upper, pred_lower = get_bounds_z(pred)
        upper_extra, upper_missed, lower_extra, lower_missed = compute_boundary_errors(
            gt_upper, gt_lower, pred_upper, pred_lower
        )
        if (
            upper_extra is not None
            and upper_missed is not None
            and lower_extra is not None
            and lower_missed is not None
        ):
            valid_rows += 1

        rows.append(
            [
                case_id,
                gt_upper if gt_upper is not None else "",
                gt_lower if gt_lower is not None else "",
                pred_upper if pred_upper is not None else "",
                pred_lower if pred_lower is not None else "",
                upper_extra if upper_extra is not None else "",
                upper_missed if upper_missed is not None else "",
                lower_extra if lower_extra is not None else "",
                lower_missed if lower_missed is not None else "",
            ]
        )

    if model_name in wb.sheetnames:
        wb.remove(wb[model_name])
    ws = wb.create_sheet(model_name)

    ws.append(columns)
    for r in rows:
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    auto_fit_and_center(ws)

    total_upper_extra = 0
    total_upper_missed = 0
    total_lower_extra = 0
    total_lower_missed = 0
    for r in rows:
        if isinstance(r[5], int):
            total_upper_extra += r[5]
        if isinstance(r[6], int):
            total_upper_missed += r[6]
        if isinstance(r[7], int):
            total_lower_extra += r[7]
        if isinstance(r[8], int):
            total_lower_missed += r[8]

    return {
        "cases_total": len(rows),
        "cases_valid": valid_rows,
        "upper_extra_total": total_upper_extra,
        "upper_missed_total": total_upper_missed,
        "lower_extra_total": total_lower_extra,
        "lower_missed_total": total_lower_missed,
    }


def write_summary_total_sheet_first(wb: openpyxl.Workbook, summary_by_model: Dict[str, Dict[str, int]]) -> None:
    if "Summary_Total" in wb.sheetnames:
        wb.remove(wb["Summary_Total"])
    ws = wb.create_sheet("Summary_Total", 0)
    ws.append(
        [
            "Model",
            "Cases Total",
            "Cases Valid",
            "Upper Extra Total",
            "Upper Missed Total",
            "Lower Extra Total",
            "Lower Missed Total",
        ]
    )

    for model_name, stats in summary_by_model.items():
        ws.append(
            [
                model_name,
                stats["cases_total"],
                stats["cases_valid"],
                stats["upper_extra_total"],
                stats["upper_missed_total"],
                stats["lower_extra_total"],
                stats["lower_missed_total"],
            ]
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)
    auto_fit_and_center(ws)


def main() -> None:
    base_dir = r"C:\Users\dell\Desktop\Rectal_146"
    gt_dir = os.path.join(base_dir, "labelsTs")
    output_excel = os.path.join(base_dir, "Eval_boundary_layer_errors.xlsx")

    model_paths = {
        "nnUNet_baseline": os.path.join(base_dir, "nnUNet_all_volume"),
        "nnUNet_post": os.path.join(base_dir, "nnUNet_all_postprocess"),
        "nnUNet_pre": os.path.join(base_dir, "nnUNet_crop_restore"),
    }

    gt_index = build_numeric_index(gt_dir)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    evaluated = 0
    summary_by_model: Dict[str, Dict[str, int]] = {}
    for model_name, pred_dir in model_paths.items():
        if not os.path.isdir(pred_dir):
            print(f"[Skip Model] Folder not found: {pred_dir}")
            continue
        summary_by_model[model_name] = evaluate_one_model(wb, model_name, pred_dir, gt_index)
        evaluated += 1

    if evaluated == 0:
        print("No model evaluated. Excel not generated.")
        return

    write_summary_total_sheet_first(wb, summary_by_model)
    wb.save(output_excel)
    print(f"\nAll done. Saved to: {output_excel}")


if __name__ == "__main__":
    main()
