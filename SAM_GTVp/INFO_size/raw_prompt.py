"""
统计GTVp上下界框和面积前三层  + 每一层面积
id
area
box_mm
xc,yc
"""

import os
import numpy as np
import pandas as pd
import SimpleITK as sitk
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# 计算所有mask切片层的id, area, box_wh, 中心点坐标xy
def get_slice_info(mask_np, spacing_x, spacing_y):
    info_list = []
    for z in range(mask_np.shape[0]):
        mask_slice = mask_np[z]
        area_pixel = np.sum(mask_slice > 0)
        if area_pixel == 0:
            continue
        area_mm2 = area_pixel * spacing_x * spacing_y

        y_indices, x_indices = np.where(mask_slice > 0)
        if len(x_indices) == 0 or len(y_indices) == 0:
            continue
        x_min = np.min(x_indices)
        x_max = np.max(x_indices)
        y_min = np.min(y_indices)
        y_max = np.max(y_indices)

        x_mid_pixel = (x_min + x_max) / 2.0
        y_mid_pixel = (y_min + y_max) / 2.0

        xc_mm = x_mid_pixel * spacing_x
        yc_mm = y_mid_pixel * spacing_y

        width_pixel = x_max - x_min + 1
        height_pixel = y_max - y_min + 1

        width_mm = width_pixel * spacing_x
        height_mm = height_pixel * spacing_y

        info_list.append((z, area_mm2, [width_mm, height_mm], (xc_mm, yc_mm)))

    return info_list

# 提取上下界和面积大小前3层
def extract_start_fields(info_list):
    if not info_list:
        return[""] * 20

    info_list.sort(key=lambda x: x[0])   # 按z大小升序排列
    up = info_list[0]
    down = info_list[-1]
    info_list.sort(key=lambda x: x[1], reverse=True)   # 按area大小降序排列
    top3 = info_list[:3]       # 取面积大小前三

    all_infos = [up, down] + top3
    flat_fields = []
    for x in all_infos:
        flat_fields += [
            x[0],
            round(x[1], 2),
            x[2],  # 保留原格式 list
            x[3],  # 保留原格式 tuple
        ]
    return flat_fields

def write_to_excel(patient_ids, data_rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SliceInfo"

    # 列组和字段名
    positions = ["上界", "下界", "面积top1", "面积top2", "面积top3"]
    subfields = ["z", "area(mm2)", "wh(mm)", "xy"]
    header_1 = ["ID"] + ["Num_Slices"] + [pos for pos in positions for _ in subfields]
    header_2 = [""] + [""] + subfields * len(positions)

    # 写入合并标题行
    ws.append(header_1)
    ws.append(header_2)

    # 合并第一行的列组（从第2列开始）
    for i in range(len(positions)):
        start_col = 3 + i * 4
        end_col = start_col + 3
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center", vertical="center")

    # 合并 PatientID
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).value = "ID"
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # 合并Num_Slices
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(row=1, column=2).value = "Num_slices"
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for i, pid in enumerate(patient_ids):
        num_slices, row_data = data_rows[i]
        flat_row = []
        for field in row_data:
            if isinstance(field, list):   # [w,h]
                flat_row.append(f"[{field[0]:.1f}, {field[1]:.1f}]")
            elif isinstance(field, tuple):   # (x,y)
                flat_row.append(f"({field[0]:.1f}, {field[1]:.1f})")
            else:
                flat_row.append(field)
        ws.append([pid, num_slices] + flat_row)

    # 居中所有单元格
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自动列宽（可选）
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)
    print(f"[完成] 已保存为 Excel：{output_path}")

# 面积表
def write_area_table(patient_ids, area_dicts, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "AreaMap"

    # 统计最大层数（用于统一列数）
    max_slices = max(len(area_dict) for area_dict in area_dicts)
    header = ["ID"] + [f"slice={i+1}" for i in range(max_slices)]
    ws.append(header)

    for pid, area_dict in zip(patient_ids, area_dicts):
        row = [pid]
        for i in range(max_slices):
            area = area_dict.get(i+1, "")
            if area != "":
                row.append(round(area, 2))  # 保留两位小数
        ws.append(row)

    # 居中对齐
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)
    print(f"[完成] 每层面积已保存为 Excel：{output_path}")




if __name__ == "__main__":
    nii_dir = r"C:\Users\dell\Desktop\SAM\GTVp_CTonly\20250809\datanii\train_nii" # nii数据文件夹 nii.gz
    output_path = r"C:/Users/dell/Desktop/train_slice_info.xlsx"
    area_output_path = r"C:/Users/dell/Desktop/train_slice_area_map.xlsx"

    patient_ids = []
    data_rows = []
    area_dicts = []

    for pa in sorted(os.listdir(nii_dir), key=lambda x: int(x.lstrip("p_"))):
        pa_mask_dir = os.path.join(nii_dir, pa)
        image_nii_path = os.path.join(nii_dir, pa, "image.nii.gz")
        # print(image_nii_path)

        # 读取 NIfTI 图像
        ct_img = sitk.ReadImage(image_nii_path)
        spacing = ct_img.GetSpacing()  # (spacing_x, spacing_y, spacing_z)
        spacing_x, spacing_y = spacing[0], spacing[1]

        # 读取mask3Dnii
        mask_nii_path = os.path.join(nii_dir, pa, "GTVp.nii.gz")
        mask_img = sitk.ReadImage(mask_nii_path)
        mask_np = sitk.GetArrayFromImage(mask_img)  # shape: (Z, H, W)
        spacing_mask = mask_img.GetSpacing()
        spacing_mask_x, spacing_mask_y = spacing_mask[0], spacing_mask[1]

        # 检查 image 和 mask 的 spacing 是否一致
        if not (spacing_x == spacing_mask_x and spacing_y == spacing_mask_y):
            print(f"{pa} spacing 不一致")
            continue

        info_list = get_slice_info(mask_np, spacing_x, spacing_y)
        z_list = sorted([x[0] for x in info_list], reverse=True)  # 将降序排列
        z_mapping = {z: i + 1 for i, z in enumerate(z_list)}     # 映射成1开始编号（靠头方向/上界）
        # print(f"{pa} z_mapping: {z_mapping}")

        info_list_renumbered= [
            (z_mapping[z], area, wh, xy)
            for (z, area, wh, xy) in info_list
        ]

        # 构造 z_idx → area 映射字典
        area_dict = {z_mapping[z]: area for (z, area, _, _) in info_list}
        area_dicts.append(area_dict)

        flat_fields = extract_start_fields(info_list_renumbered)
        n_slices = len(info_list_renumbered)   # 含mask的层数
        patient_ids.append(pa)
        data_rows.append((n_slices, flat_fields))

    write_to_excel(patient_ids, data_rows, output_path)
    write_area_table(patient_ids, area_dicts, area_output_path)












